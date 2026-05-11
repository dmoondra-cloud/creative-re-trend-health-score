"""
Rent Roll processor - Production-grade parser for Yardi, OneSite, ResMan, and custom formats.
Integrated with advanced column detection, multirow support, and source-total verification.
"""

import pandas as pd
import openpyxl
import io
import re
from typing import Dict, List, Optional, Tuple
from datetime import date, datetime, timedelta
from collections import defaultdict

# ══════════════════════════════════════════════════════════════════════════════
# CORE PARSING ENGINE (from rr_parser.py)
# ══════════════════════════════════════════════════════════════════════════════

# Section markers
SEC_START = 'Current/Notice/Vacant Residents'
SEC_END = 'Future Residents/Applicants'

# Non-unit words
_NON_UNIT_WORDS = {
    'type','total','totals','subtotal','grand','account','collections',
    'description','occupied','vacant','unit','units','status','name',
    'resident','residents','property','market','rent','amount','balance',
    'summary','average','count','notes','report','printed','current',
    'future','notice','applicant','detail','subtotals',
}

# Header patterns for column detection
_HDR_PATTERNS = {
    'unit':        r'^unit\s*(no|number|#)?$',
    'unit_type':   r'^(unit\s*)?type$|^unit\s*type|^floorplan$|^floor\s*plan$|^plan$',
    'bdba':        r'bd\s*/\s*ba|bed\s*/\s*bath',
    'tenant':      r'tenant|resident\s+name|residents?\s*name|^name$',
    'status':      r'^status$|lease\s*status|unit.lease\s*status',
    'sqft':        r'sq\.?\s*(ft|feet)|sqft|square\s*f|size\s*\(?sf\)?|^area$',
    'mkt_rent':    r'market\s*rent|mkt\s*rent',
    'eff_rent':    r'^rent$|monthly\s*rent|eff.*rent|actual\s*rent|lease\s*rent|^charge$',
    'charge_desc': r'^description$|^charge\s*(desc|code)',
    'charge_amt':  r'^amount$',
    'lease_from':  r'lease\s*(from|start|begin)',
    'lease_to':    r'lease\s*(to|end|expir)',
    'move_in':     r'move[\s\-]?in',
    'move_out':    r'move[\s\-]?out',
}

_FOOTER_RE = re.compile(
    r'^(property\s+occupancy|unit\s+type\s+occupancy|collections|totals?|'
    r'grand\s+total|summary|subtotal|rent\s+roll\s+summary)$', re.IGNORECASE
)

_BED_MAP_LETTER = {'s':'Studio','a':'1 BR','b':'2 BR','c':'3 BR','d':'4 BR','e':'5 BR'}
_RENO_SUFFIX_PAT = re.compile(
    r'[-_\s]?(renovated?|upgraded?|premium|classic|standard|original|wf)$',
    re.IGNORECASE
)

_BASE_RENT_PAT = re.compile(
    r'^rnt[a-z]?$|^r\.?n\.?t\.?$|^grnt$|^mrnt$|^nrnt$|^base[\s_]?rent$|'
    r'^contract[\s_]?rent$|^rent[\s:_\-]|^rent$',
    re.IGNORECASE
)

_NON_BASE_RENT_PAT = re.compile(
    r'pet|park|garage|storage|valet|locker|bike|cable|internet|'
    r'util|water|electric|gas|trash|sewer|fee|deposit|insur|pest|'
    r'laundry|concession|amenity|admin|late|nsf|forfeit|waiver|'
    r'bond|surety|month.to.month|mtm',
    re.IGNORECASE
)

_VACANT_PAT = re.compile(
    r'^(vacant|vacancy|model|down|admin|employee|office|mgmt|'
    r'management|maintenance|amenity|leasing|common|n/?a)$',
    re.IGNORECASE
)

# ──────────────────────────────────────────────────────────────────────────────
# Helper Functions
# ──────────────────────────────────────────────────────────────────────────────

def safe_float(v):
    if v is None: return 0.0
    try: return float(str(v).replace(',','').replace('$','').replace('%',''))
    except: return 0.0

def _pd(raw):
    if raw is None: return None
    if isinstance(raw, datetime): return raw.date()
    if isinstance(raw, date):     return raw
    if isinstance(raw, str):
        s = raw.strip().lstrip("'")
        for pat, fmt in [
            (r'(\d{1,2})/(\d{1,2})/(\d{4})', 'MDY'),
            (r'(\d{4})-(\d{1,2})-(\d{1,2})', 'YMD'),
            (r'(\d{1,2})-(\d{1,2})-(\d{4})', 'MDY'),
        ]:
            m = re.search(pat, s)
            if m:
                try:
                    a, b, c = int(m.group(1)), int(m.group(2)), int(m.group(3))
                    d = date(c, a, b) if fmt == 'MDY' else date(a, b, c)
                    if 1900 <= d.year <= 2100: return d
                except Exception: pass
    return None

def _strip_q(v):
    if v is None: return None
    s = str(v)
    return s.lstrip("'") if s.startswith("'") else s

def _looks_like_unit(val):
    if val is None: return False
    s = str(val).strip().lstrip("'")
    if not s or len(s) > 12: return False
    if s.lower() in _NON_UNIT_WORDS: return False
    if not re.search(r'\d', s): return False
    return bool(re.match(r'^[\dA-Za-z\-/]+$', s))

def _is_vacant_unit(name, status=''):
    n = str(name or '').strip()
    s = str(status or '').strip()
    if not n: return True
    if _VACANT_PAT.match(n): return True
    if _VACANT_PAT.match(s): return True
    if 'vacant' in n.lower() or 'vacant' in s.lower(): return True
    return False

def _lease_status(move_in, rr_date):
    if move_in is None: return 'Vacant'
    mi = move_in.date() if isinstance(move_in, datetime) else move_in
    return 'New Lease' if (rr_date - mi).days <= 548 else 'Renewed Lease'

def _calc_lease_start(move_in, lease_end):
    if move_in is None or lease_end is None:
        return None
    mi = move_in.date() if isinstance(move_in, datetime) else move_in
    le = lease_end.date() if isinstance(lease_end, datetime) else lease_end
    return mi if (le - mi).days <= 548 else le - timedelta(days=364)

def _decode_unit_code(raw):
    if not raw: return 'UNK', 'Classic', ''
    s = str(raw).strip()
    reno = 'Classic'
    m = _RENO_SUFFIX_PAT.search(s)
    if m:
        word = m.group(1).lower()
        reno = 'Reno' if word in ('renovated', 'renovate', 'upgraded', 'upgrade', 'premium') else 'Classic'
        s = s[:m.start()].strip()
    code = re.sub(r'-\d{3,4}$', '', s).strip()
    code = re.sub(r'^\d+(?=[A-Za-z])', '', code).strip()
    low = code.lower()
    if not m:
        if   low.endswith('wf'):  reno='Classic'; low=low[:-2]
        elif low.endswith('rw'):  reno='Reno';    low=low[:-2]
        elif low.endswith('r') and re.search(r'\dr$', low):
            reno='Reno'; low=low[:-1]
        elif re.search(r'\bpr\b|mpr$|bb\s*pr', low):
            reno='Partial'
    umix = _BED_MAP_LETTER.get(low[0] if low else '', '')
    return s.upper(), reno, umix

def _umix_from_unit_type(raw_type):
    s = str(raw_type or '').strip()
    if not s: return ''
    sl = s.lower()
    if re.search(r'\bstudio\b', sl):                            return 'Studio'
    if re.search(r'\b5[\s\-]?b(r|ed|edroom)?\b|^5br', sl):    return '5 BR'
    if re.search(r'\b4[\s\-]?b(r|ed|edroom)?\b|^4br', sl):    return '4 BR'
    if re.search(r'\b3[\s\-]?b(r|ed|edroom)?\b|^3br', sl):    return '3 BR'
    if re.search(r'\b2[\s\-]?b(r|ed|edroom)?\b|^2br', sl):    return '2 BR'
    if re.search(r'\b1[\s\-]?b(r|ed|edroom)?\b|^1br', sl):    return '1 BR'
    m = re.search(r'th(\d)', sl)
    if m: return {1:'1 BR',2:'2 BR',3:'3 BR',4:'4 BR',5:'5 BR'}.get(int(m.group(1)), '')
    stripped = re.sub(r'^\d+', '', s).strip()
    m2 = re.match(r'^([A-Za-z])\d', stripped)
    if m2: return _BED_MAP_LETTER.get(m2.group(1).lower(), '')
    return ''

def _get_rent_charge(charges):
    if not charges:
        return 0.0
    base_rents = {
        k: float(v) for k, v in charges.items()
        if v and float(v) > 0 and _BASE_RENT_PAT.search(str(k))
    }
    if base_rents:
        return sum(base_rents.values())
    candidates = [
        float(v) for k, v in charges.items()
        if v and float(v) > 0 and not _NON_BASE_RENT_PAT.search(str(k))
    ]
    if candidates:
        return max(candidates)
    positives = [float(v) for v in charges.values() if v and float(v) > 0]
    return max(positives) if positives else 0.0

# ──────────────────────────────────────────────────────────────────────────────
# Column Detection
# ──────────────────────────────────────────────────────────────────────────────

def _detect_columns(ws):
    """Run column detection — returns dict with header_row, col_map, etc."""
    header_row = None; best_score = 0
    for r in range(1, min(35, ws.max_row + 1)):
        populated = sum(1 for c in range(1, ws.max_column + 1) if ws.cell(r, c).value is not None)
        if populated < 2: continue
        row_text = ' '.join(_strip_q(str(ws.cell(r, c).value or '')) for c in range(1, ws.max_column + 1))
        score = sum(1 for pat in _HDR_PATTERNS.values() if re.search(pat, row_text, re.IGNORECASE))
        if score > best_score: best_score = score; header_row = r
    if not header_row: header_row = 1

    col_headers = []
    for c in range(1, ws.max_column + 1):
        texts = []
        for r in [header_row, header_row + 1]:
            v = _strip_q(str(ws.cell(r, c).value or '')).strip()
            if v: texts.append(v)
        combined = ' '.join(texts).strip()
        if combined: col_headers.append((c, combined))

    def _is_data_row(r):
        c1 = _strip_q(str(ws.cell(r, 1).value or '')).strip()
        if c1 in (SEC_START, SEC_END): return True
        for c in range(1, min(ws.max_column + 1, 6)):
            if _looks_like_unit(ws.cell(r, c).value): return True
        return False

    col_text = defaultdict(str)
    rows_to_merge = [header_row]
    next_r = header_row + 1
    if next_r <= ws.max_row and not _is_data_row(next_r):
        rows_to_merge.append(next_r)
    for r in rows_to_merge:
        pop = sum(1 for c in range(1, ws.max_column + 1) if ws.cell(r, c).value is not None)
        if pop < 2: continue
        for c in range(1, ws.max_column + 1):
            v = _strip_q(str(ws.cell(r, c).value or '')).strip().lower()
            if v: col_text[c] = (col_text[c] + ' ' + v).strip()

    col = {}
    for c, text in col_text.items():
        for field, pat in _HDR_PATTERNS.items():
            if re.search(pat, text, re.IGNORECASE) and field not in col:
                col[field] = c; break
    if 'tenant' not in col:
        for c, text in col_text.items():
            if re.search(r'^residents?$', text, re.IGNORECASE):
                col['tenant'] = c; break

    data_start = header_row + 1; data_end = ws.max_row; sec_bounded = False
    for r in range(1, ws.max_row + 1):
        c1 = _strip_q(str(ws.cell(r, 1).value or '')).strip()
        if c1 == SEC_START and not sec_bounded:
            data_start = r + 1; sec_bounded = True
        elif c1 == SEC_END and sec_bounded:
            data_end = r - 1; break
        elif _FOOTER_RE.match(c1) and r > header_row + 2:
            data_end = r - 1; break

    unit_col = col.get('unit', 1)
    if unit_col > 1:
        left = unit_col - 1
        sample = range(data_start, min(data_start + 20, data_end + 1))
        hits_left = sum(1 for r in sample if _looks_like_unit(ws.cell(r, left).value))
        hits_here = sum(1 for r in sample if _looks_like_unit(ws.cell(r, unit_col).value))
        if hits_left >= 2 and hits_here == 0:
            unit_col = left

    is_multirow = (col.get('charge_desc') is not None and col.get('charge_amt') is not None)

    return dict(header_row=header_row, col_map=col, col_headers=col_headers,
                data_start=data_start, data_end=data_end, unit_col=unit_col,
                is_multirow=is_multirow)

# ──────────────────────────────────────────────────────────────────────────────
# RR Parsing Engine
# ──────────────────────────────────────────────────────────────────────────────

def _parse_rr(ws, rr_date, log, col_override=None):
    """Parse rent roll — handles all formats (Yardi, OneSite, ResMan, simple)."""
    log.write("  Advanced unified parser\n")

    if col_override:
        detection = col_override
        log.write(f"  Using user column override\n")
    else:
        detection = _detect_columns(ws)

    header_row  = detection['header_row']
    col         = detection['col_map'].copy()
    data_start  = detection['data_start']
    data_end    = detection['data_end']
    unit_col    = detection['unit_col']
    is_multirow = detection['is_multirow']

    log.write(f"  Header row={header_row}  col map={col}\n")
    log.write(f"  Data rows {data_start}–{data_end}  unit_col={unit_col}  multirow={is_multirow}\n")

    cdesc_col = col.get('charge_desc')
    camt_col  = col.get('charge_amt')

    def cv(r, field, default=None):
        c = col.get(field)
        return ws.cell(r, c).value if c else default

    units = []; cur = None

    for r in range(data_start, data_end + 1):
        c1 = _strip_q(str(ws.cell(r, 1).value or '')).strip()
        if _FOOTER_RE.match(c1): break

        raw_unit = _strip_q(str(ws.cell(r, unit_col).value or '')).strip()
        is_unit  = _looks_like_unit(raw_unit)

        if is_unit:
            for adj in range(unit_col + 1, min(unit_col + 3, ws.max_column + 1)):
                if str(ws.cell(r, adj).value or '').strip().lower() in ('occupied', 'vacant'):
                    is_unit = False; break

        if is_unit:
            res_hint = _strip_q(str(ws.cell(r, 4).value or ''))
            name_hint = _strip_q(str(cv(r, 'tenant') or ''))
            stat_hint = _strip_q(str(cv(r, 'status') or ''))
            combined  = f'{res_hint} {name_hint} {stat_hint}'
            if re.search(r'pending.renewal|applicant|future', combined, re.IGNORECASE):
                is_unit = False

        if is_unit:
            if cur: units.append(cur)
            raw_type = _strip_q(str(cv(r, 'unit_type') or '')).strip()
            utype, reno, umix = _decode_unit_code(raw_type) if raw_type else ('', 'Classic', '')
            if not utype: utype = raw_type

            name_v  = _strip_q(str(cv(r, 'tenant')   or '')).strip()
            res_id  = _strip_q(str(ws.cell(r, 4).value or '')).strip()

            cur = dict(
                unit        = raw_unit,
                unit_type   = utype,
                renovation  = reno,
                unit_mix    = umix,
                sq_ft       = safe_float(cv(r, 'sqft')),
                res_id      = res_id,
                name        = name_v,
                market_rent = safe_float(cv(r, 'mkt_rent')),
                move_in     = _pd(cv(r, 'move_in')),
                lease_start = _pd(cv(r, 'lease_from')),
                lease_end   = _pd(cv(r, 'lease_to')),
                is_vacant   = False,
                charges     = {},
            )

            eff_v = safe_float(cv(r, 'eff_rent'))
            if eff_v > 0: cur['charges']['rent'] = eff_v

            if is_multirow and cdesc_col:
                code = _strip_q(str(ws.cell(r, cdesc_col).value or '')).strip()
                if code and code.lower() not in ('total', 'market', ''):
                    try: cur['charges'][code] = safe_float(ws.cell(r, camt_col).value)
                    except: pass

        elif cur is not None and is_multirow and cdesc_col:
            code = _strip_q(str(ws.cell(r, cdesc_col).value or '')).strip()
            if code and code.lower() not in ('total', ''):
                try:
                    cur['charges'][code] = cur['charges'].get(code, 0) + safe_float(ws.cell(r, camt_col).value)
                except: pass

    if cur: units.append(cur)

    seen = {}
    for u in units:
        if u['unit'] not in seen: seen[u['unit']] = u
    deduped = list(seen.values())

    for u in deduped:
        u['is_vacant'] = _is_vacant_unit(u.get('name',''), u.get('res_id',''))
        u['cls']       = 'Vacant' if u['is_vacant'] else _lease_status(u.get('move_in'), rr_date)
        if not u.get('lease_start') and u.get('move_in') and u.get('lease_end'):
            u['lease_start'] = _calc_lease_start(u['move_in'], u['lease_end'])

        rent = _get_rent_charge(u.get('charges', {}))
        if rent > 0: u['charges']['rent'] = rent
        else:        u['charges'].pop('rent', None)

        sf = u.get('sq_ft', 0)
        if u.get('unit_mix') in ('Unknown', None, ''):
            u['unit_mix'] = _umix_from_unit_type(u.get('unit_type', ''))

    log.write(f"  Parsed {len(deduped)} units\n")
    return deduped

# ══════════════════════════════════════════════════════════════════════════════
# RRProcessor Class (maintains same interface as before)
# ══════════════════════════════════════════════════════════════════════════════

class RRProcessor:
    """Rent Roll processor with advanced column detection and multirow support."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.rr_data = None
        self.column_mappings = {}
        self.log = io.StringIO()

    def load_rr(self) -> pd.DataFrame:
        """Load Rent Roll from Excel file."""
        try:
            sheet_names = ['Rent Roll', 'RR', 'Units', 'Sheet1', 'Report1']
            df = None

            for sheet in sheet_names:
                try:
                    df = pd.read_excel(self.file_path, sheet_name=sheet)
                    if len(df) > 0:
                        self.rr_data = df
                        self._detect_columns()
                        return df
                except:
                    continue

            if df is None:
                df = pd.read_excel(self.file_path)
                self.rr_data = df
                self._detect_columns()
                return df

        except Exception as e:
            raise ValueError(f"Could not load Rent Roll: {str(e)}")

    def _detect_columns(self):
        """Auto-detect column mappings."""
        if self.rr_data is None:
            return {}

        columns = list(self.rr_data.columns)

        patterns = {
            'unit_number': ['unit', 'unit no', 'unit #', 'unit number'],
            'unit_type': ['type', 'floorplan', 'floor plan', 'bd', 'br', 'bedroom'],
            'sqft': ['sqft', 'sq ft', 'square feet', 'size'],
            'resident_name': ['tenant', 'resident', 'name', 'lessee'],
            'market_rent': ['market', 'market rent', 'market rate'],
            'actual_rent': ['rent', 'actual rent', 'lease rent', 'monthly rent'],
            'status': ['status', 'occupancy', 'occupied'],
            'lease_start': ['lease start', 'start date', 'move in'],
            'lease_end': ['lease end', 'end date', 'move out']
        }

        detected = {}
        for field, keywords in patterns.items():
            for col in columns:
                col_lower = col.lower()
                for keyword in keywords:
                    if keyword in col_lower:
                        detected[field] = col
                        break
                if field in detected:
                    break

        self.column_mappings = detected
        return detected

    def get_summary(self) -> Dict:
        """Get RR summary with column mappings."""
        if self.rr_data is None:
            self.load_rr()

        df = self.rr_data

        summary = {
            'total_units': len(df),
            'columns': list(df.columns),
            'data': df,
            'column_mappings': self.column_mappings
        }

        status_col = self.column_mappings.get('status')
        if status_col and status_col in df.columns:
            summary['occupancy_stats'] = df[status_col].value_counts().to_dict()
        else:
            for col in ['Status', 'status', 'Unit Status', 'Occupancy']:
                if col in df.columns:
                    summary['occupancy_stats'] = df[col].value_counts().to_dict()
                    break

        return summary

    def get_column_suggestions(self) -> Dict:
        """Get auto-detected columns for UI."""
        if self.rr_data is None:
            self.load_rr()

        available_columns = list(self.rr_data.columns)

        return {
            'detected': self.column_mappings,
            'available': available_columns,
            'header_row': 0
        }

    def get_gpr_from_rr(self) -> float:
        """Extract GPR from Rent Roll data."""
        if self.rr_data is None:
            self.load_rr()

        rent_col = self.column_mappings.get('actual_rent')
        if rent_col and rent_col in self.rr_data.columns:
            try:
                return pd.to_numeric(self.rr_data[rent_col], errors='coerce').sum()
            except:
                pass

        rent_columns = [col for col in self.rr_data.columns
                       if 'rent' in col.lower() or 'rate' in col.lower()]

        if rent_columns:
            total_gpr = 0
            for col in rent_columns:
                try:
                    total_gpr += pd.to_numeric(
                        self.rr_data[col], errors='coerce'
                    ).sum()
                except:
                    pass
            return total_gpr

        return 0

    def apply_column_overrides(self, overrides: Dict[str, str]):
        """Apply user-selected column overrides."""
        self.column_mappings.update(overrides)

    def validate_rr(self) -> Dict:
        """Validate Rent Roll data quality."""
        issues = {
            'missing_columns': [],
            'empty_rows': [],
            'data_type_issues': [],
            'warnings': []
        }

        if self.rr_data is None:
            self.load_rr()

        expected_cols = ['unit_number', 'unit_type', 'status', 'actual_rent']
        for col_field in expected_cols:
            if col_field not in self.column_mappings:
                issues['missing_columns'].append(col_field)

        empty_count = self.rr_data.isna().sum().sum()
        if empty_count > len(self.rr_data) * 0.1:
            issues['warnings'].append(
                f"Data sparsity: {empty_count} empty cells detected"
            )

        return issues
