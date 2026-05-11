"""
Format-agnostic T12 financial statement parser.
Handles multiple formats (Yardi, MRI, etc.) and extracts line items
regardless of document structure.
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Tuple, Optional
import re
from datetime import datetime


class T12Parser:
    """Parses raw T12 files in multiple formats."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.raw_data = None
        self.property_name = None
        self.as_of_date = None
        self.months = []
        self.line_items = []

    def parse(self) -> Dict:
        """Main parsing orchestrator."""
        self._load_file()
        self._extract_metadata()
        self._extract_line_items()
        return self._format_output()

    def _load_file(self):
        """Load Excel file and detect format."""
        try:
            # Try reading with openpyxl first to detect structure
            wb = openpyxl.load_workbook(self.file_path)

            # Find the data sheet (usually first non-empty sheet)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row > 10:  # Likely has data
                    self.sheet_name = sheet_name
                    break

            # Also load with pandas for easier data handling
            self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            self.wb = wb
            self.ws = wb[self.sheet_name]

        except Exception as e:
            raise ValueError(f"Failed to load T12 file: {str(e)}")

    def _extract_metadata(self):
        """Extract property name, as-of-date, and month headers."""

        # Extract property name (look for common patterns)
        property_patterns = [
            r'Properties?:\s*(.+?)(?:\s*-|$)',
            r'Property[:\s]+(.+?)(?:\s*-|$)',
            r'^\s*(.+?)(?:\s*-\s*\d+\s*[NS][EW]|\s*\n)',
        ]

        for row in range(1, min(15, self.ws.max_row)):
            for col in range(1, 4):
                cell_val = self.ws.cell(row, col).value
                if cell_val and isinstance(cell_val, str):
                    for pattern in property_patterns:
                        match = re.search(pattern, str(cell_val), re.IGNORECASE)
                        if match and len(match.group(1)) > 3:
                            self.property_name = match.group(1).strip()
                            break
                    if self.property_name:
                        break
            if self.property_name:
                break

        # Extract as-of-date (look for date patterns)
        date_patterns = [
            r'as\s+of\s+([a-zA-Z]+\s+\d{1,2},?\s*\d{4})',
            r'period.*?([a-zA-Z]+\s+\d{1,2}\s*-?\s*[a-zA-Z]+\s+\d{1,2},?\s*\d{4})',
            r'(\d{1,2}/\d{1,2}/\d{4})',
        ]

        for row in range(1, min(12, self.ws.max_row)):
            for col in range(1, 5):
                cell_val = self.ws.cell(row, col).value
                if cell_val and isinstance(cell_val, str):
                    for pattern in date_patterns:
                        match = re.search(pattern, str(cell_val), re.IGNORECASE)
                        if match:
                            try:
                                # Parse the found date
                                date_str = match.group(1)
                                if '/' in date_str:
                                    self.as_of_date = pd.to_datetime(date_str).strftime('%m/%d/%Y')
                                else:
                                    self.as_of_date = pd.to_datetime(date_str).strftime('%m/%d/%Y')
                                break
                            except:
                                pass
                    if self.as_of_date:
                        break
            if self.as_of_date:
                break

        # Extract month headers (find row with month names)
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        for row in range(1, self.ws.max_row):
            month_count = 0
            for col in range(2, 20):
                cell_val = self.ws.cell(row, col).value
                if cell_val and isinstance(cell_val, str):
                    if any(month in str(cell_val) for month in month_names):
                        month_count += 1

            if month_count >= 3:  # Found month header row
                self.month_row = row
                for col in range(2, 20):
                    cell_val = self.ws.cell(row, col).value
                    if cell_val and isinstance(cell_val, str) and any(
                        month in str(cell_val) for month in month_names
                    ):
                        self.months.append(str(cell_val).strip())
                break

    def _extract_line_items(self):
        """Extract line items with their values, handling hierarchical structure."""

        if not self.months:
            # Fallback: use columns after first column with data
            self.months = [f'Month_{i}' for i in range(1, 13)]

        month_col_start = 2

        # Find data rows (skip headers and metadata)
        start_row = max(self.month_row + 1, 12) if hasattr(self, 'month_row') else 12

        for row in range(start_row, self.ws.max_row + 1):
            # Check for line item label in column A or B
            label_cell = self.ws.cell(row, 1).value or self.ws.cell(row, 2).value

            if not label_cell or not isinstance(label_cell, str):
                continue

            label = label_cell.strip()

            # Skip empty lines and total rows (but keep them for reference)
            if not label or label.startswith('='):
                continue

            # Determine indent level (measure leading spaces)
            indent_level = len(label) - len(label.lstrip())

            # Extract values for each month
            values = []
            for col in range(month_col_start, month_col_start + len(self.months)):
                try:
                    val = self.ws.cell(row, col).value
                    if val and isinstance(val, (int, float)):
                        values.append(float(val))
                    elif val and isinstance(val, str):
                        try:
                            values.append(float(val))
                        except:
                            values.append(0)
                    else:
                        values.append(0)
                except:
                    values.append(0)

            # Store line item
            self.line_items.append({
                'label': label.strip(),
                'indent': indent_level,
                'values': values,
                'row': row,
                'is_subtotal': 'total' in label.lower(),
                'is_section_header': indent_level < 4 and 'total' not in label.lower()
            })

    def _format_output(self) -> Dict:
        """Format parsed data for use in categorization."""
        return {
            'property_name': self.property_name or 'Unknown Property',
            'as_of_date': self.as_of_date or '',
            'months': self.months,
            'line_items': self.line_items,
            'parsed_successfully': len(self.line_items) > 0
        }

    def get_line_items_by_category(self, category: str) -> List[Dict]:
        """Get all line items under a specific category (Income/Expense)."""
        return [
            item for item in self.line_items
            if item['is_section_header'] or (
                category.lower() in str(item['label']).lower()
            )
        ]
