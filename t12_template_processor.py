"""
T12 Template processor - populates categorized data into Creative RE's T12 template.
Handles insertion of line items, preservation of formulas, and metadata updates.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy
from typing import Dict, List, Tuple
import re


class T12TemplateProcessor:
    """Populates categorized T12 data into the template."""

    def __init__(self, template_path: str):
        self.template_path = template_path
        self.wb = openpyxl.load_workbook(template_path)
        self.ws = self.wb['T12']

    def populate_t12(self,
                    categorized_items: List[Dict],
                    property_name: str,
                    as_of_date: str,
                    months: List[str]) -> openpyxl.Workbook:
        """
        Populate categorized T12 data into template.
        Returns modified workbook.
        """

        # Update metadata
        self._update_metadata(property_name, as_of_date)

        # Update months if needed
        self._update_months(months)

        # Populate categorized line items
        self._populate_line_items(categorized_items)

        return self.wb

    def _update_metadata(self, property_name: str, as_of_date: str):
        """Update property name and as-of-date in T12 sheet."""
        try:
            # Property name in C4
            self.ws['C4'] = property_name

            # As-of date in C6
            if as_of_date:
                self.ws['C6'] = as_of_date

        except Exception as e:
            print(f"Warning: Could not update metadata: {str(e)}")

    def _update_months(self, months: List[str]):
        """Update month headers if provided."""
        try:
            # Months typically start in column E (month 1)
            for idx, month in enumerate(months[:12]):  # Max 12 months
                col_letter = self._get_column_letter(5 + idx)  # E is column 5
                self.ws[f'{col_letter}8'] = month
        except Exception as e:
            print(f"Warning: Could not update months: {str(e)}")

    def _populate_line_items(self, categorized_items: List[Dict]):
        """
        Populate line items starting from row 9 (data starts after headers).
        Matches items to their respective categories in template.
        """

        # Group items by category/section
        items_by_section = {}
        for item in categorized_items:
            section = item.get('section', 'Other')
            if section not in items_by_section:
                items_by_section[section] = []
            items_by_section[section].append(item)

        # Find section headers in template and insert data
        current_row = 9
        section_positions = {}

        # Scan template to find where sections are
        for row in range(9, self.ws.max_row + 1):
            cell_value = self.ws.cell(row, 2).value  # Column B usually has labels
            if cell_value and isinstance(cell_value, str):
                for section in items_by_section.keys():
                    if section.lower() in cell_value.lower():
                        section_positions[section] = row
                        break

        # Populate items under their sections
        for section, items in items_by_section.items():
            if section in section_positions:
                insert_row = section_positions[section] + 1
                self._insert_items_at_row(items, insert_row)

    def _insert_items_at_row(self, items: List[Dict], start_row: int):
        """Insert line items starting at a specific row."""
        try:
            for idx, item in enumerate(items):
                row = start_row + idx

                # Insert row if needed
                if row > self.ws.max_row:
                    self.ws.insert_rows(row)

                # Label in column D (usually where "Particulars" goes)
                self.ws.cell(row, 4).value = item['label']

                # Category in column C
                self.ws.cell(row, 3).value = item.get('category', '')

                # Values in columns E onwards (months)
                values = item.get('values', [])
                for col_idx, value in enumerate(values[:12]):
                    col_letter = self._get_column_letter(5 + col_idx)
                    self.ws[f'{col_letter}{row}'].value = value

                # Add formula in column B for total (SUM of months)
                col_letters = [self._get_column_letter(5 + i) for i in range(len(values))]
                if col_letters:
                    formula = f"=SUM({col_letters[0]}{row}:{col_letters[-1]}{row})"
                    self.ws.cell(row, 2).value = formula

        except Exception as e:
            print(f"Warning: Could not insert items at row {start_row}: {str(e)}")

    def _copy_cell_style(self, source_cell, target_cell):
        """Copy formatting from one cell to another."""
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    @staticmethod
    def _get_column_letter(col_num: int) -> str:
        """Convert column number to letter."""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + (col_num % 26)) + result
            col_num //= 26
        return result

    def preserve_validations(self):
        """
        Preserve any data validations in the template.
        Important: User specified not to remove validations.
        """
        # Data validations are already preserved when loading with openpyxl
        # This method is here for explicit documentation
        pass

    def save(self, output_path: str):
        """Save populated workbook to file."""
        try:
            self.wb.save(output_path)
            return True
        except Exception as e:
            print(f"Error saving T12: {str(e)}")
            return False


class T12Validator:
    """Validates populated T12 data."""

    @staticmethod
    def validate_t12(wb: openpyxl.Workbook) -> Dict:
        """Run validation checks on populated T12."""
        issues = {
            'missing_property_name': False,
            'missing_dates': False,
            'formula_errors': [],
            'missing_categories': [],
            'warnings': []
        }

        ws = wb['T12']

        # Check metadata
        if not ws['C4'].value:
            issues['missing_property_name'] = True

        if not ws['C6'].value:
            issues['missing_dates'] = True

        # Check for formulas with errors
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula cell
                    try:
                        # Check if formula references exist
                        if '#REF!' in str(cell.value):
                            issues['formula_errors'].append(f"Row {cell.row}, Col {cell.column}")
                    except:
                        pass

        return issues
