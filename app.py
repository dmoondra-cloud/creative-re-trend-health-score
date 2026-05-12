"""
Creative RE — UW Pipeline  (Streamlit web app)
================================================
Upload:  (1) THS Template with T12 already categorised
         (2) Raw Rent Roll

Output:  Completed Trend Health Scorecard — ready to download

Deploy:  streamlit run app.py   (local)
         OR push to GitHub → connect at share.streamlit.io
"""

import io, re, zipfile
from copy import copy as _copy
from datetime import date, datetime, timedelta
from collections import defaultdict

import streamlit as st
import openpyxl
import requests as _requests
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _xls_to_xlsx_bytes(xls_bytes):
    """
    Convert a legacy .xls file (xlrd) to .xlsx bytes (openpyxl).
    Returns xlsx bytes, or raises if xlrd is not installed.
    """
    try:
        import xlrd
    except ImportError:
        raise ImportError("xlrd is required to read .xls files. Add xlrd to requirements.txt.")

    book = xlrd.open_workbook(file_contents=xls_bytes)
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for sheet_idx in range(book.nsheets):
        src  = book.sheet_by_index(sheet_idx)
        dst  = wb.create_sheet(title=src.name)
        for row in range(src.nrows):
            for col in range(src.ncols):
                cell = src.cell(row, col)
                val  = cell.value
                # xlrd types: 0=empty,1=text,2=number,3=date,4=bool,5=error
                if cell.ctype == 3 and val:
                    try:
                        val = xlrd.xldate_as_datetime(val, book.datemode)
                    except Exception:
                        pass
                elif cell.ctype == 0:
                    val = None
                dst.cell(row + 1, col + 1, val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
#  COLUMN MAP  (0-indexed) — adjust if your export differs
# ══════════════════════════════════════════════════════════════════════════════
COL_UNIT=0; COL_UTYPE=1; COL_SF=2; COL_RESID=3; COL_NAME=4; COL_MKT=5
COL_CHARGE=6; COL_AMT=7; COL_MOVEIN=8; COL_LEASEEND=9; COL_MOVEOUT=10
COL_RESDEP=11; COL_OTHDEP=12; COL_BAL=13

SEC_START = 'Current/Notice/Vacant Residents'
SEC_END   = 'Future Residents/Applicants'

# ══════════════════════════════════════════════════════════════════════════════
#  T12 CATEGORY SETS
# ══════════════════════════════════════════════════════════════════════════════
_GPR_CATS  = {"Gross Potential Rents"}
_TEI_CATS  = {"Gross Potential Rents","Less: Loss to Lease","Less: Vacancy Loss",
              "Less: Non-Revenue Units","Less: Concessions","Less: Bad Debt"}
_CONC_CATS = {"Less: Concessions"}
_BD_CATS   = {"Less: Bad Debt"}
_OTHER_CATS = {"RUBS","Late Fee/NSF/Termination Fee","Pet Rent","Renter Insurance",
               "Fee Income","Miscellaneous Revenue","Garage Income",
               "Cable Commissions","Laundry","Pest Control Income",
               "Legal or Attorney Charges"}
_EXP_CATS  = {"Management Fees","Personnel Costs","Administrative",
              "Advertising & Promotion","Repairs & Maintenance","Turnover",
              "Contract Services","Landscaping","Utilities",
              "Franchise Tax","Real Estate Taxes","Insurance","Capital Reserves"}

# ══════════════════════════════════════════════════════════════════════════════
#  STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════
C_NAVY="1F3A5F"; C_DGRAY="363735"; C_META="D6E4F0"; C_ALT="EFF5FB"
C_GREEN="E2EFDA"; C_YELLOW="FFF2CC"; C_RED="FCE4D6"; C_WHITE="FFFFFF"
C_TXT="1A1A1A"; C_GRNDK="375623"; C_YELDK="7F6000"; C_REDDK="C00000"
APPROX_TOL = 0.015

def _fill(h):   return PatternFill("solid", fgColor=h)
def _fnt(bold=False, sz=10, color=C_TXT, italic=False):
    return Font(name="Arial", bold=bold, size=sz, color=color, italic=italic)
def _aln(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _bdr():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, r, col, val, bg=C_WHITE, bold=False, sz=10, color=C_TXT,
       nf=None, ha='left', va='center', wrap=False, italic=False):
    cl = ws.cell(row=r, column=col, value=val)
    cl.fill=_fill(bg); cl.font=_fnt(bold,sz,color,italic)
    cl.alignment=_aln(ha,va,wrap); cl.border=_bdr()
    if nf: cl.number_format=nf
    return cl

def cas(ws, r, c1, c2, val, bg, bold=True, sz=11, color=C_WHITE, h=22):
    ws.row_dimensions[r].height = h
    for col in range(c1, c2+1):
        cl = ws.cell(row=r, column=col, value=val if col==c1 else None)
        cl.fill=_fill(bg); cl.font=_fnt(bold,sz,color)
        cl.alignment=Alignment(horizontal='centerContinuous', vertical='center')
        cl.border=_bdr()

def safe_float(v):
    if v is None: return 0.0
    try: return float(str(v).replace(',','').replace('$','').replace('%',''))
    except: return 0.0

def _pd(raw):
    if raw is None: return None
    if isinstance(raw, datetime): return raw.date()
    if isinstance(raw, date):     return raw
    # Handle string dates (e.g. from OneSite/RealPage .xls exports)
    if isinstance(raw, str):
        s = raw.strip().lstrip("'")
        for pat, fmt in [
            (r'(\d{1,2})/(\d{1,2})/(\d{4})', 'MDY'),
            (r'(\d{4})-(\d{1,2})-(\d{1,2})', 'YMD'),
            (r'(\d{1,2})-(\d{1,2})-(\d{4})', 'MDY'),
        ]:
            import re as _re
            m = _re.search(pat, s)
            if m:
                try:
                    a, b, c = int(m.group(1)), int(m.group(2)), int(m.group(3))
                    d = date(c, a, b) if fmt == 'MDY' else date(a, b, c)
                    if 1900 <= d.year <= 2100: return d
                except Exception: pass
    return None

# ══════════════════════════════════════════════════════════════════════════════
#  RENT ROLL PARSER  — auto-detects format:
#
#  Format A (simple one-liner):  one row per unit, headers = Unit|BD/BA|Tenant|Status|...
#  Format B (multi-row detail):  multiple rows per unit with separate charge-code rows
#                                and section markers (Current/Notice/Vacant Residents)
# ══════════════════════════════════════════════════════════════════════════════

def _beds_from_bdba(bdba):
    """'3/2.00' → '3 BR', '1/1.00' → '1 BR', '0/1.00' → 'Studio'."""
    try:
        beds = int(str(bdba).split('/')[0])
        return {0:'Studio',1:'1 BR',2:'2 BR',3:'3 BR',4:'4 BR',5:'5 BR'}.get(beds, f'{beds} BR')
    except: return str(bdba)

def _unit_type_from_bdba(bdba):
    """'3/2.00' → '3BR/2BA'"""
    try:
        parts = str(bdba).split('/')
        beds  = int(parts[0])
        baths = parts[1].rstrip('0').rstrip('.') if len(parts) > 1 else '?'
        return f'{beds}BR/{baths}BA'
    except: return str(bdba)

_BED_MAP_LETTER = {'s':'Studio','a':'1 BR','b':'2 BR','c':'3 BR','d':'4 BR','e':'5 BR'}

# Words appended after the bedroom letter that indicate renovation level
_RENO_SUFFIX_PAT = re.compile(
    r'[-_\s]?(renovated?|upgraded?|premium|classic|standard|original|wf)$',
    re.IGNORECASE
)

def _decode_unit_code(raw):
    """
    Decode letter-based unit codes like A2R-872, 5493TH3R, sf1992c2-Classic.
    Returns (unit_type_str, renovation_label, unit_mix_label).
    """
    if not raw: return 'UNK', 'Classic', ''
    s = str(raw).strip()

    # Detect renovation from word suffix (-Classic, -Renovated, -Upgraded, etc.)
    reno = 'Classic'
    m = _RENO_SUFFIX_PAT.search(s)
    if m:
        word = m.group(1).lower()
        reno = 'Reno' if word in ('renovated', 'renovate', 'upgraded', 'upgrade', 'premium') else 'Classic'
        s = s[:m.start()].strip()   # strip the reno word from the code

    # Strip numeric property prefix (e.g. "5493" in "5493TH3R", "872" in "A2R-872")
    # Pattern: leading digits followed by letters, or trailing "-digits"
    code = re.sub(r'-\d{3,4}$', '', s).strip()          # strip trailing -NNN
    code = re.sub(r'^\d+(?=[A-Za-z])', '', code).strip() # strip leading digits before letter

    low = code.lower()

    # Detect renovation from letter suffix R / WF (after stripping word suffix).
    # Only treat R as Renovated when it is a STANDALONE suffix directly after
    # the base type code (digit + single R), e.g. A1R, B2R, C1r.
    # Multi-letter suffixes like MPR, PR, BBR, BB PR are property-specific
    # designations — NOT renovation indicators — so we leave them as Classic.
    if not m:
        if   low.endswith('wf'):  reno='Classic'; low=low[:-2]
        elif low.endswith('rw'):  reno='Reno';    low=low[:-2]
        elif low.endswith('r') and re.search(r'\dr$', low):
            # R is Reno only when it directly follows a digit: a1r, b2r, th3r → Reno
            # Multi-letter suffixes like MPR, PR, BB PR → digit is NOT before final R → Classic
            reno='Reno'; low=low[:-1]
        elif re.search(r'\bpr\b|mpr$|bb\s*pr', low):
            # PR / MPR / BB PR = Partial renovation
            reno='Partial'

    # Bedroom map from first letter of remaining code
    umix = _BED_MAP_LETTER.get(low[0] if low else '', '')

    return s.upper(), reno, umix

def _lease_status(move_in, rr_date):
    if move_in is None: return 'Vacant'
    mi = move_in.date() if isinstance(move_in, datetime) else move_in
    return 'New Lease' if (rr_date - mi).days <= 548 else 'Renewed Lease'

_VACANT_PAT = re.compile(
    r'^(vacant|vacancy|model|down|admin|employee|office|mgmt|'
    r'management|maintenance|amenity|leasing|common|n/?a)$',
    re.IGNORECASE
)

def _is_vacant_unit(name, status=''):
    """
    Determine vacancy from the resident name and/or status fields directly.
    A unit is vacant when the name is empty, explicitly says 'VACANT', or
    matches another known non-revenue label (Model, Down, Admin, etc.).
    No tenant ID parsing needed.
    """
    n = str(name or '').strip()
    s = str(status or '').strip()
    if not n: return True
    if _VACANT_PAT.match(n): return True
    if _VACANT_PAT.match(s): return True
    if 'vacant' in n.lower() or 'vacant' in s.lower(): return True
    return False

def _calc_lease_start(move_in, lease_end):
    """
    Calculate lease start when not provided directly on the RR:
      - If (lease_end − move_in) ≤ 18 months (548 days) → lease_start = move_in
      - Else → lease_start = lease_end − 364 days
    Both inputs should be date objects; returns a date or None.
    """
    if move_in is None or lease_end is None:
        return None
    mi = move_in.date() if isinstance(move_in, datetime) else move_in
    le = lease_end.date() if isinstance(lease_end, datetime) else lease_end
    return mi if (le - mi).days <= 548 else le - timedelta(days=364)

# ── Rent charge detection ─────────────────────────────────────────────────────
# PRIMARY: unambiguous base-rent codes/labels — match these first, exclusively.
_BASE_RENT_PAT = re.compile(
    r'^rnt[a-z]?$'                    # rntn, rntr, rnt, rntm …
    r'|^r\.?n\.?t\.?$'                # R.N.T.
    r'|^grnt$|^mrnt$|^nrnt$'          # other PMS shortcodes
    r'|^base[\s_]?rent$'
    r'|^contract[\s_]?rent$'
    r'|^rent[\s:_\-]'                 # "Rent: Resident", "Rent_charge" etc.
    r'|^rent$',                        # exactly "rent"
    re.IGNORECASE
)

# NON-RENT: charges that look like "_____ Rent" but are NOT base rent
_NON_BASE_RENT_PAT = re.compile(
    r'pet|park|garage|storage|valet|locker|bike|cable|internet|'
    r'util|water|electric|gas|trash|sewer|fee|deposit|insur|pest|'
    r'laundry|concession|amenity|admin|late|nsf|forfeit|waiver|'
    r'bond|surety|month.to.month|mtm',
    re.IGNORECASE
)

def _get_rent_charge(charges):
    """
    Extract the effective base rent from a charges dict (code/description → amount).
    Strategy:
      1. Find charges matching primary base-rent patterns (rnta, 'Rent: Resident', etc.)
         → sum those, ignoring anything else. This prevents 'Pet Rent', 'Garage Rent'
         etc. from inflating the total just because they contain the word 'rent'.
      2. If nothing matched, fall back to the largest positive charge that doesn't
         look like a non-rent fee.
      3. Last resort: largest positive charge of any kind.
    """
    if not charges:
        return 0.0

    # Step 1 — unambiguous base-rent codes/descriptions only
    base_rents = {
        k: float(v) for k, v in charges.items()
        if v and float(v) > 0 and _BASE_RENT_PAT.search(str(k))
    }
    if base_rents:
        return sum(base_rents.values())

    # Step 2 — largest positive charge that doesn't look like a fee/addon
    candidates = [
        float(v) for k, v in charges.items()
        if v and float(v) > 0 and not _NON_BASE_RENT_PAT.search(str(k))
    ]
    if candidates:
        return max(candidates)

    # Step 3 — absolute fallback: largest positive charge
    positives = [float(v) for v in charges.values() if v and float(v) > 0]
    return max(positives) if positives else 0.0


# Header keyword patterns — maps field names to regex patterns for column detection
_HDR_PATTERNS = {
    'unit':        r'^unit\s*(no|number|#)?$',
    'unit_type':   r'^(unit\s*)?type$|^unit\s*type|^floorplan$|^floor\s*plan$|^plan$',
    'bdba':        r'bd\s*/\s*ba|bed\s*/\s*bath',
    'tenant':      r'tenant|resident\s+name|residents?\s*name|^name$',
    'status':      r'^status$|lease\s*status|unit.lease\s*status',
    'sqft':        r'sq\.?\s*(ft|feet)|sqft|square\s*f|size\s*\(?sf\)?|^area$',
    'mkt_rent':    r'market\s*rent|mkt\s*rent',
    'eff_rent':    r'^rent$|monthly\s*rent|eff.*rent|actual\s*rent|lease\s*rent|^charge$',
    'charge_desc': r'^description$|^charge\s*(desc|code)',
    'charge_amt':  r'^amount$',
    'lease_from':  r'lease\s*(from|start|begin)',
    'lease_to':    r'lease\s*(to|end|expir)',
    'move_in':     r'move[\s\-]?in',
    'move_out':    r'move[\s\-]?out',
}

_FLOOR_PLAN_HEADERS = {
    'Studio': r'studio',
    '1 BR':   r'1[\s\-]?bed(?:room)?|one[\s\-]?bed(?:room)?',
    '2 BR':   r'2[\s\-]?bed(?:room)?|two[\s\-]?bed(?:room)?',
    '3 BR':   r'3[\s\-]?bed(?:room)?|three[\s\-]?bed(?:room)?',
    '4 BR':   r'4[\s\-]?bed(?:room)?|four[\s\-]?bed(?:room)?',
    '5 BR':   r'5[\s\-]?bed(?:room)?|five[\s\-]?bed(?:room)?',
}

def _parse_floor_plan_text(text):
    """
    Parse raw page text for patterns like '1 Bed · 650-850 sq ft'.
    Returns list of (sf_min, sf_max, unit_mix_label).
    """
    results = []
    for label, bed_pat in _FLOOR_PLAN_HEADERS.items():
        chunks = re.findall(rf'.{{0,120}}(?:{bed_pat}).{{0,120}}', text, re.IGNORECASE)
        for chunk in chunks:
            # Range: "650 - 850" or "650–850"
            m = re.search(r'(\d{3,4})\s*[-–to]+\s*(\d{3,4})', chunk)
            if m:
                lo, hi = int(m.group(1)), int(m.group(2))
                if 200 < lo < hi < 5000:
                    results.append((lo, hi, label))
            else:
                # Single value near "sq ft"
                m2 = re.search(r'(\d{3,4})\s*sq\.?\s*ft', chunk, re.IGNORECASE)
                if m2:
                    sf = int(m2.group(1))
                    if 200 < sf < 5000:
                        results.append((sf - 30, sf + 30, label))
    return results

def _lookup_unit_mix_online(prop_name, address_hint='', direct_url='', log=None):
    """
    Fetch floor plan data online.
    Search order:
      0. Direct URL pasted by user (most reliable — fetch immediately)
      1. DuckDuckGo HTML search
      2. Property website extracted from DuckDuckGo results
      3. Apartments.com search
    Returns list of (sf_min, sf_max, unit_mix_label), or [].
    Never raises.
    """
    _HDR = {
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36'
        ),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
    }
    query = f"{prop_name} {address_hint}".strip()

    def _fetch_text(url):
        try:
            r = _requests.get(url, headers=_HDR, timeout=12, allow_redirects=True)
            if r.status_code == 200:
                t = re.sub(r'<[^>]+>', ' ', r.text)
                return re.sub(r'\s+', ' ', t)
        except Exception:
            pass
        return ''

    def _try(url, src):
        text = _fetch_text(url)
        plans = _parse_floor_plan_text(text)
        if log:
            log.write(f"  [{src}] {len(plans)} floor plan range(s) found\n" if plans
                      else f"  [{src}] no floor plan data\n")
        return plans

    # 0 — User-pasted URL (property website or Apartments.com listing)
    if direct_url and direct_url.startswith('http'):
        plans = _try(direct_url, f'user URL: {direct_url[:60]}')
        if plans:
            return plans

    # 1 — DuckDuckGo HTML
    ddg_url = (f"https://html.duckduckgo.com/html/?q="
               f"{_requests.utils.quote(query + ' apartments floor plans sq ft bedroom')}")
    plans = _try(ddg_url, 'DuckDuckGo')
    if plans:
        return plans

    # 2 — First few real URLs extracted from DuckDuckGo results
    try:
        from urllib.parse import unquote
        ddg_text = _fetch_text(ddg_url)
        for enc in re.findall(r'uddg=([^&"]+)', ddg_text)[:3]:
            try:
                site_url = unquote(enc)
                if any(s in site_url for s in ('duckduckgo', 'google', 'bing')):
                    continue
                plans = _try(site_url, f'site:{site_url[:50]}')
                if plans:
                    return plans
            except Exception:
                continue
    except Exception:
        pass

    # 3 — Apartments.com search
    apt_url = f"https://www.apartments.com/search/?q={_requests.utils.quote(query)}"
    plans = _try(apt_url, 'apartments.com')
    if plans:
        return plans

    if log:
        log.write("  ⚠ No floor plan data found — unit mix left blank for ambiguous units\n")
    return []

def _apply_floor_plans(plans, sq_ft):
    """
    Map a unit's sq ft to a unit mix label using fetched floor plan ranges.
    1. Exact range match (lo ≤ sq_ft ≤ hi).
    2. Closest mid-point match — used when sq_ft sits just outside all ranges
       (e.g. website shows 850–1050 but unit is 1060 due to rounding).
    Returns '' if plans is empty or sq_ft is falsy.
    """
    if not plans or not sq_ft:
        return ''
    # Exact match (also handles single-value entries like lo==hi with ±5% tolerance)
    for lo, hi, label in plans:
        effective_lo = lo * 0.95 if lo == hi else lo
        effective_hi = hi * 1.05 if lo == hi else hi
        if effective_lo <= sq_ft <= effective_hi:
            return label
    # Closest mid-point match (within 20 % of range mid to avoid wild mismatches)
    best_label, best_dist = '', float('inf')
    for lo, hi, label in plans:
        mid  = (lo + hi) / 2
        dist = abs(sq_ft - mid)
        span = max(hi - lo, 1)
        if dist < best_dist and dist / span < 1.5:   # within 1.5× the range width
            best_dist, best_label = dist, label
    return best_label


_AS_OF_PAT = re.compile(
    r'as[\s\-]?of|report\s*date|through\s*date|period\s*end|printed\b',
    re.IGNORECASE
)

def _parse_date_from_string(s):
    """Extract a valid 2020–2030 date from a string. Returns date or None."""
    for pat, fmt in [
        (r'(\d{1,2})/(\d{1,2})/(\d{4})', 'MDY'),
        (r'(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})', 'YMD'),
        (r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', 'MDY'),
        (r'([A-Za-z]+)\s+(\d{1,2}),?\s+(\d{4})', 'MonDY'),
    ]:
        m = re.search(pat, s)
        if not m: continue
        try:
            if fmt == 'MDY':
                d = date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
            elif fmt == 'YMD':
                d = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            else:  # MonDY: "March 3, 2026"
                d = datetime.strptime(f"{m.group(1)} {m.group(2)} {m.group(3)}", '%B %d %Y').date()
            if 2020 <= d.year <= 2030:
                return d
        except Exception:
            pass
    return None


def _extract_date_from_rr_file(file_bytes, log):
    """
    Scan the top 25 rows of the RR file looking for an 'as of' label and
    extracting the date from that cell or the cell(s) immediately adjacent.
    Never picks up tenant lease dates — only cells labelled as the report date.
    Returns a date object or None.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb['Report1'] if 'Report1' in wb.sheetnames else wb.worksheets[0]
        max_col = min(ws.max_column, 20)

        for r in range(1, min(25, ws.max_row + 1)):
            for c in range(1, max_col + 1):
                val = ws.cell(r, c).value
                if val is None: continue
                s = str(val).strip()

                # Case 1: cell contains "as of <date>" all in one string
                # e.g. "As of 03/03/2026" or "Report Date: 2026-03-03"
                if _AS_OF_PAT.search(s):
                    d = _parse_date_from_string(s)
                    if d:
                        if log: log.write(f"  RR date from 'as of' cell ({r},{c}): {d}\n")
                        return d
                    # Date may be in the next cell(s) on the same row
                    for nc in range(c + 1, min(c + 4, max_col + 1)):
                        nval = ws.cell(r, nc).value
                        if nval is None: continue
                        if isinstance(nval, (date, datetime)):
                            d = nval.date() if isinstance(nval, datetime) else nval
                            if 2020 <= d.year <= 2030:
                                if log: log.write(f"  RR date from cell after 'as of' ({r},{nc}): {d}\n")
                                return d
                        nd = _parse_date_from_string(str(nval))
                        if nd:
                            if log: log.write(f"  RR date from cell after 'as of' ({r},{nc}): {nd}\n")
                            return nd
                    # Date may be in the cell below
                    bval = ws.cell(r + 1, c).value
                    if bval is not None:
                        if isinstance(bval, (date, datetime)):
                            d = bval.date() if isinstance(bval, datetime) else bval
                            if 2020 <= d.year <= 2030:
                                if log: log.write(f"  RR date from cell below 'as of' ({r+1},{c}): {d}\n")
                                return d
                        nd = _parse_date_from_string(str(bval))
                        if nd:
                            if log: log.write(f"  RR date from cell below 'as of' ({r+1},{c}): {nd}\n")
                            return nd

                # Case 2: bare date/datetime cell — only trust if row has a label nearby
                if isinstance(val, (date, datetime)):
                    d = val.date() if isinstance(val, datetime) else val
                    if 2020 <= d.year <= 2030:
                        row_text = ' '.join(
                            str(ws.cell(r, cc).value or '') for cc in range(1, max_col + 1)
                        )
                        if _AS_OF_PAT.search(row_text):
                            if log: log.write(f"  RR date (bare date in labelled row) ({r},{c}): {d}\n")
                            return d

        # ── Fallback: find the header row, then scan rows ABOVE it for any date ──
        # Rows above the unit header are always report metadata (name, date, company),
        # never tenant lease data — so any parseable date there is the report date.
        header_row = None
        for r in range(1, min(30, ws.max_row + 1)):
            for c in range(1, max_col + 1):
                if re.match(r'^unit\s*(no|number|#)?$',
                            str(ws.cell(r, c).value or '').strip(), re.IGNORECASE):
                    header_row = r
                    break
            if header_row:
                break

        scan_to = (header_row - 1) if header_row else min(10, ws.max_row)
        for r in range(1, scan_to + 1):
            for c in range(1, max_col + 1):
                val = ws.cell(r, c).value
                if val is None: continue
                # Try to parse any date from cell text
                d = _parse_date_from_string(str(val))
                if d:
                    if log: log.write(f"  RR date from pre-header row {r}: {d}\n")
                    return d

    except Exception:
        pass
    return None


def _extract_date_from_t12(ths_bytes, log):
    """
    Read the last populated month column header from the T12 sheet to infer the as-of date.
    T12 column headers are typically in row 7, cols E-P (months oldest→newest).
    Returns a date object or None.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(ths_bytes), data_only=True)
        if 'T12' not in wb.sheetnames:
            return None
        ws = wb['T12']
        # Scan row 7 cols E-P (5-16) for the last non-empty cell that looks like a date
        last_date = None
        for c in range(5, 17):
            val = ws.cell(7, c).value
            if val is None: continue
            if isinstance(val, (date, datetime)):
                d = val.date() if isinstance(val, datetime) else val
                if 2020 <= d.year <= 2030:
                    last_date = d
            else:
                s = str(val).strip()
                # Handle "Jan-25", "Mar-26" style headers
                m = re.match(r'([A-Za-z]{3})[\-\s](\d{2,4})$', s)
                if m:
                    try:
                        yr = int(m.group(2))
                        if yr < 100: yr += 2000
                        mo = datetime.strptime(m.group(1), '%b').month
                        import calendar
                        last_day = calendar.monthrange(yr, mo)[1]
                        last_date = date(yr, mo, last_day)
                    except Exception:
                        pass
        if last_date and log:
            log.write(f"  RR date from T12 last month: {last_date}\n")
        return last_date
    except Exception:
        return None



# Words that must never be treated as unit IDs
_NON_UNIT_WORDS = {
    'type','total','totals','subtotal','grand','account','collections',
    'description','occupied','vacant','unit','units','status','name',
    'resident','residents','property','market','rent','amount','balance',
    'summary','average','count','notes','report','printed','current',
    'future','notice','applicant','detail','subtotals',
}

def _looks_like_unit(val):
    """True if val is a plausible unit identifier (alphanumeric, contains digit)."""
    if val is None: return False
    s = str(val).strip().lstrip("'")
    if not s or len(s) > 12: return False
    if s.lower() in _NON_UNIT_WORDS: return False
    if not re.search(r'\d', s): return False
    return bool(re.match(r'^[\dA-Za-z\-/]+$', s))


def _strip_q(v):
    """Strip Yardi leading apostrophe from cell values."""
    if v is None: return None
    s = str(v)
    return s.lstrip("'") if s.startswith("'") else s


_FOOTER_RE = re.compile(
    r'^(property\s+occupancy|unit\s+type\s+occupancy|collections|totals?|'
    r'grand\s+total|summary|subtotal|rent\s+roll\s+summary)$', re.IGNORECASE
)


def _umix_from_unit_type(raw_type):
    """Infer unit mix label from a unit type code string."""
    s = str(raw_type or '').strip()
    if not s: return ''
    sl = s.lower()
    if re.search(r'\bstudio\b', sl):                            return 'Studio'
    if re.search(r'\b5[\s\-]?b(r|ed|edroom)?\b|^5br', sl):    return '5 BR'
    if re.search(r'\b4[\s\-]?b(r|ed|edroom)?\b|^4br', sl):    return '4 BR'
    if re.search(r'\b3[\s\-]?b(r|ed|edroom)?\b|^3br', sl):    return '3 BR'
    if re.search(r'\b2[\s\-]?b(r|ed|edroom)?\b|^2br', sl):    return '2 BR'
    if re.search(r'\b1[\s\-]?b(r|ed|edroom)?\b|^1br', sl):    return '1 BR'
    m = re.search(r'th(\d)', sl)
    if m: return {1:'1 BR',2:'2 BR',3:'3 BR',4:'4 BR',5:'5 BR'}.get(int(m.group(1)), '')
    stripped = re.sub(r'^\d+', '', s).strip()
    m2 = re.match(r'^([A-Za-z])\d', stripped)
    if m2: return _BED_MAP_LETTER.get(m2.group(1).lower(), '')
    return ''


def _detect_columns(ws):
    """
    Run column detection only — no row parsing.
    Returns a dict with: header_row, col_map, col_headers, data_start, unit_col, is_multirow.
    col_headers = list of (col_index, header_text) for every non-empty column in header row.
    """
    # Find header row
    header_row = None; best_score = 0
    for r in range(1, min(35, ws.max_row + 1)):
        populated = sum(1 for c in range(1, ws.max_column + 1) if ws.cell(r, c).value is not None)
        if populated < 2: continue
        row_text = ' '.join(_strip_q(str(ws.cell(r, c).value or '')) for c in range(1, ws.max_column + 1))
        score = sum(1 for pat in _HDR_PATTERNS.values() if re.search(pat, row_text, re.IGNORECASE))
        if score > best_score: best_score = score; header_row = r
    if not header_row: header_row = 1

    # Collect column headers
    col_headers = []
    for c in range(1, ws.max_column + 1):
        texts = []
        for r in [header_row, header_row + 1]:
            v = _strip_q(str(ws.cell(r, c).value or '')).strip()
            if v: texts.append(v)
        combined = ' '.join(texts).strip()
        if combined: col_headers.append((c, combined))

    # Build column map
    def _is_data_row(r):
        c1 = _strip_q(str(ws.cell(r, 1).value or '')).strip()
        if c1 in (SEC_START, SEC_END): return True
        for c in range(1, min(ws.max_column + 1, 6)):
            if _looks_like_unit(ws.cell(r, c).value): return True
        return False

    col_text = defaultdict(str)
    rows_to_merge = [header_row]
    next_r = header_row + 1
    if next_r <= ws.max_row and not _is_data_row(next_r):
        rows_to_merge.append(next_r)
    for r in rows_to_merge:
        pop = sum(1 for c in range(1, ws.max_column + 1) if ws.cell(r, c).value is not None)
        if pop < 2: continue
        for c in range(1, ws.max_column + 1):
            v = _strip_q(str(ws.cell(r, c).value or '')).strip().lower()
            if v: col_text[c] = (col_text[c] + ' ' + v).strip()

    col = {}
    for c, text in col_text.items():
        for field, pat in _HDR_PATTERNS.items():
            if re.search(pat, text, re.IGNORECASE) and field not in col:
                col[field] = c; break
    if 'tenant' not in col:
        for c, text in col_text.items():
            if re.search(r'^residents?$', text, re.IGNORECASE):
                col['tenant'] = c; break

    # Data boundaries
    data_start = header_row + 1; data_end = ws.max_row; sec_bounded = False
    for r in range(1, ws.max_row + 1):
        c1 = _strip_q(str(ws.cell(r, 1).value or '')).strip()
        if c1 == SEC_START and not sec_bounded:
            data_start = r + 1; sec_bounded = True
        elif c1 == SEC_END and sec_bounded:
            data_end = r - 1; break
        elif _FOOTER_RE.match(c1) and r > header_row + 2:
            data_end = r - 1; break

    # Unit column (may be 1 left of header)
    unit_col = col.get('unit', 1)
    if unit_col > 1:
        left = unit_col - 1
        sample = range(data_start, min(data_start + 20, data_end + 1))
        hits_left = sum(1 for r in sample if _looks_like_unit(ws.cell(r, left).value))
        hits_here = sum(1 for r in sample if _looks_like_unit(ws.cell(r, unit_col).value))
        if hits_left >= 2 and hits_here == 0:
            unit_col = left

    is_multirow = (col.get('charge_desc') is not None and col.get('charge_amt') is not None)

    return dict(header_row=header_row, col_map=col, col_headers=col_headers,
                data_start=data_start, data_end=data_end, unit_col=unit_col,
                is_multirow=is_multirow)


def _parse_rr(ws, rr_date, log, floor_plans=None, col_override=None):
    """
    Single unified RR parser. Handles all formats:
    - Yardi / OneSite detail  (section markers, multi-row charge lines)
    - ResMan multirow          (no section markers, Description + Amount cols)
    - Simple one-line          (one row per unit, all fields in header-mapped cols)
    - Any mix of the above

    Steps:
    1  Find the header row  — most field-name matches, skip single-cell title rows
    2  Build column map     — merge up to 2 header rows, two-pass for tenant
    3  Find data boundaries — SEC_START/SEC_END markers or footer keywords
    4  Detect unit column   — may be 1 left of the 'Unit' header label
    5  Detect multirow      — presence of Description + Amount charge columns
    6  Parse every row      — unit rows vs charge rows, skip footer / pending rows
    7  Post-process         — dedup, vacancy, lease-start, rent, unit-mix
    """
    log.write("  Unified parser\n")

    # ── Steps 1-5: column detection (or use user-provided override) ──────────
    if col_override:
        # User confirmed/corrected the mapping — use it directly
        detection = col_override
        log.write(f"  Using user column override\n")
    else:
        detection = _detect_columns(ws)

    header_row  = detection['header_row']
    col         = detection['col_map'].copy()
    data_start  = detection['data_start']
    data_end    = detection['data_end']
    unit_col    = detection['unit_col']
    is_multirow = detection['is_multirow']

    log.write(f"  Header row={header_row}  col map={col}\n")
    log.write(f"  Data rows {data_start}–{data_end}  unit_col={unit_col}  multirow={is_multirow}\n")

    cdesc_col = col.get('charge_desc')
    camt_col  = col.get('charge_amt')

    # Helper: get cell value at 1-based col
    def cv(r, field, default=None):
        c = col.get(field)
        return ws.cell(r, c).value if c else default

    log.write(f"  eff_col={col.get('eff_rent')}\n")

    # ── 6. Parse rows ────────────────────────────────────────────────────────
    units = []; cur = None

    for r in range(data_start, data_end + 1):
        # Hard stop at footer keywords
        c1 = _strip_q(str(ws.cell(r, 1).value or '')).strip()
        if _FOOTER_RE.match(c1): break

        raw_unit = _strip_q(str(ws.cell(r, unit_col).value or '')).strip()
        is_unit  = _looks_like_unit(raw_unit)

        # Footer summary rows: 'Occupied'/ 'Vacant' in col immediately right of unit col
        if is_unit:
            for adj in range(unit_col + 1, min(unit_col + 3, ws.max_column + 1)):
                if str(ws.cell(r, adj).value or '').strip().lower() in ('occupied', 'vacant'):
                    is_unit = False; break

        # Skip pending renewal / applicant rows
        if is_unit:
            res_hint = _strip_q(str(ws.cell(r, 4).value or ''))  # col 4 = res_id in Yardi
            name_hint = _strip_q(str(cv(r, 'tenant') or ''))
            stat_hint = _strip_q(str(cv(r, 'status') or ''))
            combined  = f'{res_hint} {name_hint} {stat_hint}'
            if re.search(r'pending.renewal|applicant|future', combined, re.IGNORECASE):
                is_unit = False

        if is_unit:
            if cur: units.append(cur)
            raw_type = _strip_q(str(cv(r, 'unit_type') or '')).strip()
            utype, reno, umix = _decode_unit_code(raw_type) if raw_type else ('', 'Classic', '')
            if not utype: utype = raw_type

            name_v  = _strip_q(str(cv(r, 'tenant')   or '')).strip()
            res_id  = _strip_q(str(ws.cell(r, 4).value or '')).strip()  # Yardi res-id col

            cur = dict(
                unit        = raw_unit,
                unit_type   = utype,
                renovation  = reno,
                unit_mix    = umix,
                sq_ft       = safe_float(cv(r, 'sqft')),
                res_id      = res_id,
                name        = name_v,
                market_rent = safe_float(cv(r, 'mkt_rent')),
                move_in     = _pd(cv(r, 'move_in')),
                lease_start = _pd(cv(r, 'lease_from')),
                lease_end   = _pd(cv(r, 'lease_to')),
                is_vacant   = False,
                charges     = {},
            )

            # Direct effective rent column (one-row-per-unit files)
            eff_v = safe_float(cv(r, 'eff_rent'))
            if eff_v > 0: cur['charges']['rent'] = eff_v

            # First charge on same row (multirow)
            if is_multirow and cdesc_col:
                code = _strip_q(str(ws.cell(r, cdesc_col).value or '')).strip()
                if code and code.lower() not in ('total', 'market', ''):
                    try: cur['charges'][code] = safe_float(ws.cell(r, camt_col).value)
                    except: pass

        elif cur is not None and is_multirow and cdesc_col:
            code = _strip_q(str(ws.cell(r, cdesc_col).value or '')).strip()
            if code and code.lower() not in ('total', ''):
                try:
                    cur['charges'][code] = cur['charges'].get(code, 0) + safe_float(ws.cell(r, camt_col).value)
                except: pass

    if cur: units.append(cur)

    # ── 7. Post-process ──────────────────────────────────────────────────────
    seen = {}
    for u in units:
        if u['unit'] not in seen: seen[u['unit']] = u
    deduped = list(seen.values())

    for u in deduped:
        u['is_vacant'] = _is_vacant_unit(u.get('name',''), u.get('res_id',''))
        u['cls']       = 'Vacant' if u['is_vacant'] else _lease_status(u.get('move_in'), rr_date)
        if not u.get('lease_start') and u.get('move_in') and u.get('lease_end'):
            u['lease_start'] = _calc_lease_start(u['move_in'], u['lease_end'])

        rent = _get_rent_charge(u.get('charges', {}))
        if rent > 0: u['charges']['rent'] = rent
        else:        u['charges'].pop('rent', None)

        sf = u.get('sq_ft', 0)
        if floor_plans and sf:
            web_mix = _apply_floor_plans(floor_plans, sf)
            if web_mix: u['unit_mix'] = web_mix; continue
        if u.get('unit_mix') in ('Unknown', None, ''):
            u['unit_mix'] = _umix_from_unit_type(u.get('unit_type', ''))

    log.write(f"  Parsed {len(deduped)} units\n")
    return deduped


def _infer_mix_from_peers(units, log):
    """
    For units that still have no unit_mix, infer it from already-classified
    units on the same rent roll using sq ft centroid matching.

    Steps:
    1. Compute the average sq ft per unit mix label from classified units.
    2. For each unclassified unit, assign the label whose centroid is closest
       in sq ft — provided the distance is within 30 % of that centroid's value.

    This correctly classifies TH units (and any other ambiguous codes) whose
    sq ft overlaps with regular floor plans already decoded from unit type codes.
    """
    from collections import defaultdict

    # Build centroid table from classified units
    sf_by_mix = defaultdict(list)
    for u in units:
        mix = u.get('unit_mix', '')
        sf  = u.get('sq_ft', 0)
        if mix and mix not in ('', 'Unknown') and sf > 0:
            sf_by_mix[mix].append(sf)

    if not sf_by_mix:
        return

    centroids = {mix: sum(sfs)/len(sfs) for mix, sfs in sf_by_mix.items()}
    log.write(f"  SF centroids: { {m: round(c) for m,c in centroids.items()} }\n")

    fixed = 0
    for u in units:
        if u.get('unit_mix') in ('', 'Unknown', None) and u.get('sq_ft', 0) > 0:
            sf = u['sq_ft']
            best_mix, best_dist = '', float('inf')
            for mix, cent in centroids.items():
                dist = abs(sf - cent)
                if dist < best_dist:
                    best_dist, best_mix = dist, mix
            # Accept if within 30 % of the winning centroid's value
            if best_mix and best_dist / centroids[best_mix] <= 0.30:
                u['unit_mix'] = best_mix
                fixed += 1

    if fixed:
        log.write(f"  Inferred unit mix for {fixed} unclassified units from sq ft centroids\n")


def _extract_rr_source_totals(ws, log):
    """
    Extract the property's own summary totals from the RR footer.
    Looks for two specific sections common in PMS exports:
      1. A charge-summary section with a 'Rent: Resident' line → effective rent
      2. A 'Property Occupancy' section with a labelled header row and a grand
         total row that has units, market rent and square footage.
    Returns dict: {n_units, mkt, eff, sf} — None for any not found.
    """
    result = {'n_units': None, 'mkt': None, 'eff': None, 'sf': None}
    MAX_COL = min(ws.max_column, 30)

    rent_pat = re.compile(r'rent[\s:]+resident|resident[\s:]+rent', re.IGNORECASE)

    # ── Pass 1: find effective rent total ─────────────────────────────────────
    for r in range(1, ws.max_row + 1):
        for c in range(1, MAX_COL + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and rent_pat.search(v):
                # Largest numeric on this row > 10,000 is the rent total
                nums = [float(ws.cell(r, cc).value) for cc in range(1, MAX_COL + 1)
                        if isinstance(ws.cell(r, cc).value, (int, float))
                        and float(ws.cell(r, cc).value) > 10000]
                if nums:
                    result['eff'] = max(nums)
                    log.write(f"  Source eff rent R{r}: {result['eff']:,.0f}\n")
                break

    # ── Pass 2: find Property Occupancy grand total ────────────────────────────
    # Locate the 'Property Occupancy' header (skip 'Unit Type Occupancy')
    occ_row = None
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(r, 1).value or '').strip()
        if re.search(r'^property\s+occupancy$', v, re.IGNORECASE):
            occ_row = r
            break

    if occ_row:
        # Next row should have column headers — read column positions
        hdr_row = occ_row + 1
        col_mkt = col_units = col_sf = None
        for c in range(1, MAX_COL + 1):
            h = str(ws.cell(hdr_row, c).value or '').strip().lower()
            if 'market' in h:                col_mkt   = c
            elif 'unit' in h and 'type' not in h: col_units = c
            elif 'square' in h or (h.startswith('sq') and 'ft' in h): col_sf = c

        # Grand total row: comes after "Total Vacant" row — identified by having
        # ONLY numeric values (no text in any column) within the next 10 rows
        found_vacant = False
        for r in range(hdr_row + 1, min(occ_row + 15, ws.max_row + 1)):
            c4 = str(ws.cell(r, 4).value or '').strip().lower()
            if c4 == 'vacant':
                found_vacant = True
                continue
            if found_vacant:
                # First row after vacant row — check it's all numeric (grand total)
                text_vals = [ws.cell(r, c).value for c in range(1, MAX_COL + 1)
                             if isinstance(ws.cell(r, c).value, str)
                             and ws.cell(r, c).value.strip()]
                if not text_vals:  # no text cells = grand total row
                    if col_units:
                        v = ws.cell(r, col_units).value
                        if isinstance(v, (int, float)): result['n_units'] = int(v)
                    if col_mkt:
                        v = ws.cell(r, col_mkt).value
                        if isinstance(v, (int, float)): result['mkt'] = float(v)
                    if col_sf:
                        v = ws.cell(r, col_sf).value
                        if isinstance(v, (int, float)): result['sf'] = float(v)
                    log.write(f"  Source totals from Property Occupancy R{r}: "
                              f"units={result['n_units']} mkt={result['mkt']} sf={result['sf']}\n")
                    break

    # ── Fallback for Yardi detail format: scan for totals after SEC_END ───────
    if result['n_units'] is None:
        sec_end_pat = re.compile(r'future residents', re.IGNORECASE)
        past_end = False
        for r in range(1, ws.max_row + 1):
            c1 = str(ws.cell(r, 1).value or '').strip("'").strip()
            if sec_end_pat.search(c1):
                past_end = True
            if past_end:
                # Look for rows with "total" label and numeric sibling
                if re.search(r'\btotal\b', c1, re.IGNORECASE):
                    for c in range(2, MAX_COL + 1):
                        v = ws.cell(r, c).value
                        if isinstance(v, (int, float)) and v > 100:
                            # Check what kind of value this might be
                            if 10 <= v <= 2000 and result['n_units'] is None:
                                result['n_units'] = int(v)
                            elif v > 50000 and result['mkt'] is None:
                                result['mkt'] = float(v)

    log.write(f"  Source totals final: {result}\n")
    return result


def parse_rent_roll(file_bytes, rr_date, log, floor_plans=None, col_override=None):
    wb  = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    src = wb['Report1'] if 'Report1' in wb.sheetnames else wb.worksheets[0]
    log.write(f"  Sheet: '{src.title}'  ({src.max_row} rows)\n")

    # Single unified parser — handles all RR formats in one pass
    units = _parse_rr(src, rr_date, log, floor_plans, col_override=col_override)

    # Infer unit mix for any still-unclassified units using sq ft centroids
    _infer_mix_from_peers(units, log)

    # ── Computed totals (what the parser produced) ────────────────────────────
    n   = len(units)
    sf  = sum(u['sq_ft'] for u in units)
    mkt = sum(u['market_rent'] for u in units)
    eff = sum(u['charges'].get('rent', 0) for u in units)
    log.write(f"  {n} units  SF={sf:,.0f}  MktRent={mkt:,.0f}  EffRent={eff:,.0f}\n")

    # ── Source totals from RR footer (used for verification table S50:S53) ───
    # These come from the RR's own summary section so any parser discrepancy
    # is correctly flagged as an error in the template's Match? column.
    _footer = _extract_rr_source_totals(src, log)
    # Only use footer values that were actually found — never fall back to computed.
    # Unfound values stay None so the verification table leaves those cells blank
    # rather than showing a meaningless "ALL GOOD" (source == calculated trivially).
    source_totals = dict(
        n_units    = _footer['n_units'],
        mkt        = _footer['mkt'],
        eff        = _footer['eff'],
        sf         = _footer['sf'],
        _found_n   = _footer['n_units'] is not None,
        _found_mkt = _footer['mkt']     is not None,
        _found_eff = _footer['eff']     is not None,
        _found_sf  = _footer['sf']      is not None,
    )
    def _fmt(v): return f"{v:,.0f}" if v is not None else "N/A"
    log.write(f"  Verification 'From Source': units={source_totals['n_units']} "
              f"mkt={_fmt(source_totals['mkt'])} eff={_fmt(source_totals['eff'])} "
              f"sf={_fmt(source_totals['sf'])}\n")

    return units, dict(n_units=n, sf=sf, mkt=mkt, eff=eff), source_totals


# ══════════════════════════════════════════════════════════════════════════════
#  WRITE RENT ROLL TAB  (one-liner function removed — no longer used)
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
#  WRITE RENT ROLL TAB
# ══════════════════════════════════════════════════════════════════════════════
TPL_E_MAX=289; DATA_START=8

def _get_dv_unit_mix_values(ws):
    """
    Read the allowed unit mix values from the data validation on col F (unit mix column).
    Returns a list of strings, or [] if no validation found.
    """
    allowed = []
    for dv in ws.data_validations.dataValidation:
        # Check if the validation covers column F (col 6) in the data area
        for rng in dv.sqref.ranges:
            if rng.min_col <= 6 <= rng.max_col:
                # formula1 is typically '"Studio,1 Bedroom,2 Bedroom,..."'
                f = str(dv.formula1 or '').strip('"').strip("'")
                if f:
                    allowed = [v.strip() for v in f.split(',') if v.strip()]
                break
        if allowed:
            break
    return allowed


def _snap_to_dv(value, allowed):
    """
    Match value to closest allowed label (case-insensitive).
    Returns the matched allowed value, or value unchanged if no match.
    """
    if not allowed or not value:
        return value
    vl = str(value).strip().lower()
    for a in allowed:
        if a.strip().lower() == vl:
            return a          # exact match
    # Partial match — e.g. "1 bedroom" in allowed["1 Bedroom"]
    for a in allowed:
        if vl in a.lower() or a.lower() in vl:
            return a
    return value              # no match — keep as-is (Excel will show validation warning)


def write_rr_tab(wb, units, src_totals, source_totals, rr_date, prop_name, log):
    ws = wb['Rent Roll']

    # Read data-validation allowed values for unit mix (col F) from the template
    dv_mixes = _get_dv_unit_mix_values(ws)
    if dv_mixes:
        log.write(f"  Unit mix DV values: {dv_mixes}\n")

    ws.cell(4,3).value = prop_name
    ws.cell(5,3).value = rr_date

    # ── Write unit data rows ───────────────────────────────────────────────────
    last_row = DATA_START + len(units) - 1
    for i, u in enumerate(units):
        r=DATA_START+i; iv=u['is_vacant']
        ws.cell(r,2).value = u['unit']
        ws.cell(r,3).value = u['unit_type']
        ws.cell(r,4).value = u['renovation']
        ws.cell(r,6).value = _snap_to_dv(u['unit_mix'], dv_mixes)
        ws.cell(r,7).value = int(u['sq_ft']) if u.get('sq_ft') else None
        ws.cell(r,8).value = u.get('move_in')
        ws.cell(r,9).value = u.get('lease_start')
        ws.cell(r,10).value= u.get('lease_end')
        ws.cell(r,13).value= u['name']
        ws.cell(r,14).value= float(u['market_rent']) if u.get('market_rent') else None
        ws.cell(r,15).value= float(u['charges'].get('rent',0)) or None  # 0 → blank

    # ── Fix column E formula range ─────────────────────────────────────────────
    if last_row > TPL_E_MAX:
        for r in range(TPL_E_MAX+1, last_row+1):
            ws.cell(r,5).value = f'=F{r}&"-"&D{r}&"-"&C{r}'
    for r in range(last_row+1, TPL_E_MAX+1):
        ws.cell(r,5).value = None

    # ── Populate R8:R41 with unique floor-plan labels (col E values) ───────────
    # Col E = unit_mix-Renovation-UnitType (e.g. "2 Bedroom-Classic-2BR/2BA").
    # We compute this in Python (same logic as the col-E formula =F&"-"&D&"-"&C)
    # and write as plain values — no SORT/UNIQUE formula needed.
    # openpyxl cannot encode dynamic-array formulas; attempting to write them
    # causes Excel to flag the file as requiring repair.
    seen_fps: set = set()
    fp_list: list = []
    for u in units:
        mix  = u.get('unit_mix')  or ''
        reno = u.get('renovation') or ''
        utp  = u.get('unit_type')  or ''
        label = f"{mix}-{reno}-{utp}".strip('-')
        if label and label not in seen_fps:
            seen_fps.add(label)
            fp_list.append(label)
    fp_list.sort()

    for r in range(8, 42):                          # clear R8:R41
        ws.cell(r, 18).value = None
    for i, label in enumerate(fp_list):             # write sorted unique labels
        if 8 + i > 41: break                        # R8:R41 = 34 slots max
        ws.cell(8 + i, 18).value = label
    log.write(f"  R8:R{7+len(fp_list)} populated with {len(fp_list)} unique floor plans: {fp_list}\n")

    # ── Verification table: 'From Source' = RR footer totals only ─────────────
    # Only write a value when the RR's own summary section actually stated it.
    # If the summary wasn't found (None), leave the cell blank — this prevents
    # a false "ALL GOOD" that would appear if we fell back to our own computed totals.
    _pre = _extract_rr_source_totals
    ws.cell(50,19).value = source_totals['n_units'] if source_totals.get('_found_n')   else None
    ws.cell(51,19).value = source_totals['mkt']     if source_totals.get('_found_mkt') else None
    ws.cell(52,19).value = source_totals['eff']     if source_totals.get('_found_eff') else None
    ws.cell(53,19).value = source_totals['sf']      if source_totals.get('_found_sf')  else None
    log.write(f"  Rent Roll tab written ({len(units)} units, rows {DATA_START}–{last_row})\n")


# ══════════════════════════════════════════════════════════════════════════════
#  READ T12 METRICS (monthly body — stable value cells)
# ══════════════════════════════════════════════════════════════════════════════
def _ann(m):
    return dict(T12=sum(m),T9=sum(m[3:])/9*12,T6=sum(m[6:])*2,T3=sum(m[9:])*4,T1=m[-1]*12)

def read_t12_metrics(wb, log):
    ws = wb['T12']
    cat_m = defaultdict(lambda:[0.0]*12)
    rows=0
    for r in range(9, ws.max_row+1):
        cat = str(ws.cell(r,3).value or '').strip()
        if not cat or cat in ('None','Category'): continue
        vals = [safe_float(ws.cell(r,c).value) for c in range(5,17)]
        if any(v!=0 for v in vals):
            for i,v in enumerate(vals): cat_m[cat][i]+=v
            rows+=1
    if rows==0:
        raise ValueError("T12 tab has no monthly data. Make sure it is fully categorised.")

    log.write(f"  T12 rows parsed: {rows}  Categories: {len(cat_m)}\n")
    def cs(cats): return [sum(cat_m[c][i] for c in cats) for i in range(12)]
    gpr_m=cs(_GPR_CATS); tei_m=cs(_TEI_CATS); conc_m=cs(_CONC_CATS)
    bd_m=cs(_BD_CATS); oi_m=cs(_OTHER_CATS); exp_m=cs(_EXP_CATS)
    noi_m=[tei_m[i]+oi_m[i]-exp_m[i] for i in range(12)]
    tei=_ann(tei_m); oi=_ann(oi_m); noi=_ann(noi_m)
    gpr=_ann(gpr_m); conc=_ann(conc_m); bd=_ann(bd_m)
    def ratio(n,d): return {p:(n[p]/d[p] if d[p] else 0.0) for p in ('T12','T9','T6','T3','T1')}
    rg_pct=(tei['T3']-tei['T12'])/tei['T12'] if tei['T12'] else 0.0
    og_pct=(oi['T3'] -oi['T12']) /oi['T12']  if oi['T12']  else 0.0
    m = {'Rental Income':tei,'Other Income':oi,'NOI':noi,
         'Economic Occupancy (%)':ratio(tei,gpr),
         'Concessions (% of GPR)':{p:(-abs(conc[p])/gpr[p] if gpr[p] else 0.0) for p in ('T12','T9','T6','T3','T1')},
         'Bad Debt (% of GPR)':   {p:(-abs(bd[p])  /gpr[p] if gpr[p] else 0.0) for p in ('T12','T9','T6','T3','T1')},
         '_rg_pct':rg_pct,'_og_pct':og_pct}
    for r in range(57,90):
        lbl=str(ws.cell(r,25).value or '').strip()
        if 'Rental' in lbl and 'Growth' in lbl:
            v=safe_float(ws.cell(r,26).value)
            if v: m['_rg_pct']=v
        elif 'Other' in lbl and 'Growth' in lbl:
            v=safe_float(ws.cell(r,26).value)
            if v: m['_og_pct']=v
    log.write(f"  TEI T12={tei['T12']:,.0f}  T3={tei['T3']:,.0f}  "
              f"NOI T12={noi['T12']:,.0f}  EconOcc={m['Economic Occupancy (%)']['T12']*100:.1f}%\n")
    return m


# ══════════════════════════════════════════════════════════════════════════════
#  RR METRICS
# ══════════════════════════════════════════════════════════════════════════════
def read_rr_metrics(units, src, rr_date, log):
    cut90=rr_date-timedelta(days=90)
    n=len(units); nv=sum(1 for u in units if u['cls']=='Vacant'); no=n-nv
    occ=[u for u in units if u['cls']!='Vacant']
    avg_mkt=src['mkt']/n if n else 0.0
    avg_eff=src['eff']/no if no else 0.0
    ip=(avg_mkt-avg_eff)/avg_mkt if avg_mkt else 0.0
    l90=[u for u in units if u.get('lease_start') and cut90<=u['lease_start']<=rr_date]
    fp_occ=defaultdict(list); fp_rec=defaultdict(list)
    for u in occ: fp_occ[u['unit_type']].append(u['charges'].get('rent',0))
    for u in l90: fp_rec[u['unit_type']].append(u['charges'].get('rent',0))
    fps=list(fp_occ.keys())
    pos=sum(1 for fp in fps if fp_occ[fp] and fp_rec.get(fp) and
            (sum(fp_rec[fp])/len(fp_rec[fp]))/(sum(fp_occ[fp])/len(fp_occ[fp]))-1>0)
    breadth=pos/len(fps) if fps else 0.0
    new_l=[u for u in units if u['cls']=='New Lease'     and u['charges'].get('rent',0)>0]
    ren_l=[u for u in units if u['cls']=='Renewed Lease' and u['charges'].get('rent',0)>0]
    avg_new=sum(u['charges'].get('rent',0) for u in new_l)/len(new_l) if new_l else 0.0
    avg_ren=sum(u['charges'].get('rent',0) for u in ren_l)/len(ren_l) if ren_l else 0.0
    spread=(avg_new/avg_ren-1) if avg_ren else 0.0
    log.write(f"  Occ {no}/{n}  Breadth {breadth*100:.1f}%  Spread {spread*100:.2f}%  In-place vs mkt {ip*100:.2f}%\n")
    return dict(rr_date=rr_date,n_total=n,n_vacant=nv,n_occupied=no,
                avg_market=avg_mkt,avg_eff=avg_eff,in_place_spread=ip,
                breadth_pct=breadth,n_fps=len(fps),n_leases_90d=len(l90),
                n_new_leases=len(new_l),n_renewed_leases=len(ren_l),
                avg_new_eff=avg_new,avg_renewed_eff=avg_ren,tradeout_spread=spread)


# ══════════════════════════════════════════════════════════════════════════════
#  SCORING
# ══════════════════════════════════════════════════════════════════════════════
def trend_dir(t12,t6,t3):
    gt=lambda a,b: a>b*(1+APPROX_TOL)
    ap=lambda a,b: abs(a-b)<=max(abs(a),abs(b),1)*APPROX_TOL
    if gt(t3,t6) and gt(t6,t12): return 5
    if gt(t3,t6) and ap(t6,t12): return 4
    if ap(t3,t6) and ap(t6,t12): return 3
    if gt(t6,t3) and ap(t6,t12): return 2
    if gt(t6,t3) and gt(t12,t6): return 1
    return 3
def s_eoo(p):  return 5 if p>=.95 else 4 if p>=.92 else 3 if p>=.88 else 2 if p>=.85 else 1
def s_con(p):  return 5 if p<=.005 else 4 if p<=.01 else 3 if p<=.02 else 2 if p<=.03 else 1
def s_bd(p):   return 5 if p<=.005 else 4 if p<=.01 else 3 if p<=.02 else 2 if p<=.03 else 1
def s_gr(p,h5,h4,h3,h2): return 5 if p>=h5 else 4 if p>=h4 else 3 if p>=h3 else 2 if p>=h2 else 1
def s_ip(sp):  return 5 if sp>=.05 else 4 if sp>=.02 else 3 if sp>=0 else 2 if sp>=-.04 else 1

def score_all(t12, rr, log):
    ri=t12['Rental Income']; oi=t12['Other Income']; noi=t12['NOI']
    eoo=t12['Economic Occupancy (%)']['T12']
    con=t12['Concessions (% of GPR)']['T12']
    bd =t12['Bad Debt (% of GPR)']['T12']
    noi_s=trend_dir(noi['T12'],noi['T6'],noi['T3'])
    eoo_s=s_eoo(eoo); con_s=s_con(abs(con)); bd_s=s_bd(abs(bd))
    rg=t12['_rg_pct']; og=t12['_og_pct']
    rt=trend_dir(ri['T12'],ri['T6'],ri['T3']); rg_s=s_gr(rg,.03,.01,0,-.02)
    ot=trend_dir(oi['T12'],oi['T6'],oi['T3']); og_s=s_gr(og,.05,.01,-.01,-.04)
    rt_comp=.4*rt+.4*rg_s+.1*ot+.1*og_s; rt_s=max(1,min(5,round(rt_comp)))
    rt_d=dict(rt=rt,rg=rg_s,ot=ot,og=og_s,rg_pct=rg,og_pct=og,composite=rt_comp,score=rt_s)
    bs=5 if rr['breadth_pct']>=.80 else 4 if rr['breadth_pct']>=.60 else 3 if rr['breadth_pct']>=.40 else 2 if rr['breadth_pct']>=.25 else 1
    ss=5 if rr['tradeout_spread']>=.03 else 4 if rr['tradeout_spread']>=.01 else 3 if rr['tradeout_spread']>=0 else 2 if rr['tradeout_spread']>=-.02 else 1
    to_comp=.5*bs+.5*ss; to_s=max(1,min(5,round(to_comp)))
    to_d=dict(breadth_score=bs,spread_score=ss,composite=to_comp,score=to_s)
    ip_s=s_ip(rr['in_place_spread'])
    log.write(f"  NOI={noi_s}  EconOcc={eoo_s}({eoo*100:.1f}%)  Conc={con_s}  BD={bd_s}  "
              f"RevTrend={rt_s}  Tradeout={to_s}  InPlace={ip_s}\n")
    return dict(noi=noi_s,econ_occ=eoo_s,concessions=con_s,bad_debt=bd_s,
                rev_trend=rt_s,rev_trend_detail=rt_d,tradeout=to_s,tradeout_detail=to_d,inplace=ip_s)


# ══════════════════════════════════════════════════════════════════════════════
#  BUILD TREND HEALTH SCORE SHEET
# ══════════════════════════════════════════════════════════════════════════════
WEIGHTS={'noi':20,'econ_occ':15,'concessions':15,'bad_debt':15,'rev_trend':15,'tradeout':15,'inplace':5}
BANDS=[(85,'Strong'),(70,'Healthy'),(55,'Mixed / Watch'),(40,'Caution'),(0,'High Risk')]
def get_band(s):
    for t,l in BANDS:
        if s>=t: return l
    return 'High Risk'

def build_ths_sheet(wb, t12, rr, scores, today_str, prop_name, log):
    """
    Write computed scores into the existing 'Trend Health Score' template sheet.
    All original formatting, formulas, and layout are preserved — only data
    cells (property name, date, scores, direction text, notes) are updated.

    Template layout (1-based):
      R2,C3  : Property Name value
      R4,C3  : As-Of Date value
      R8–R14 : one row per metric
        col D (4): Score (1–5)  ← write here
        col E (5): =(D/5)*C    ← formula, DO NOT TOUCH
        col F (6): Trend Direction text  ← write here
        col G (7): Notes / Evidence      ← write here
      R15,E  : =SUM(E8:E14)   ← formula, DO NOT TOUCH
      R16,E  : =IF(...)        ← formula, DO NOT TOUCH
    """
    if 'Trend Health Score' not in wb.sheetnames:
        raise ValueError("THS template has no 'Trend Health Score' sheet.")
    ws = wb['Trend Health Score']

    # ── Metadata ──────────────────────────────────────────────────────────────
    ws.cell(2, 3).value = prop_name
    ws.cell(4, 3).value = today_str

    # ── Build per-metric direction + notes strings ────────────────────────────
    ps = lambda v, d=1: f"{v*100:.{d}f}%"
    ds = lambda v: (f"${v/1e6:.2f}M" if abs(v) >= 1e6
                    else f"${v/1e3:.0f}K" if abs(v) >= 1e3 else f"${v:.0f}")

    noi = t12['NOI']; ri = t12['Rental Income']; oi = t12['Other Income']
    eoo_v = t12['Economic Occupancy (%)']['T12']
    con_v = abs(t12['Concessions (% of GPR)']['T12'])
    bd_v  = abs(t12['Bad Debt (% of GPR)']['T12'])
    rv    = scores['rev_trend_detail']
    tot   = scores['tradeout_detail']
    ip    = rr['in_place_spread']
    n12, n6, n3 = noi['T12'], noi['T6'], noi['T3']

    nd = (f"T3 {ds(n3)} > T6 {ds(n6)} > T12 {ds(n12)} ↑" if n3 > n6 > n12 * 1.001 else
          f"T3 {ds(n3)} > T6 {ds(n6)} ↑ Recent"            if n3 > n6 else
          f"T3 {ds(n3)} < T6 {ds(n6)} ↓ Declining"         if n3 < n6 < n12 * .999 else
          f"T3 {ds(n3)} | T6 {ds(n6)} | T12 {ds(n12)} → Flat")

    # (score, direction, notes) for each row in order R8..R14
    metric_data = [
        (scores['noi'],
         nd,
         f"NOI T12={ds(n12)} | T6={ds(n6)} | T3={ds(n3)} | T1={ds(noi['T1'])}"),
        (scores['econ_occ'],
         f"{ps(eoo_v)} TEI/GPR (T12)",
         f"Economic Occupancy T12={ps(eoo_v)}."),
        (scores['concessions'],
         f"{ps(con_v)} of GPR (T12)",
         f"Concessions T12={ps(con_v)} of GPR."),
        (scores['bad_debt'],
         f"{ps(bd_v)} of GPR (T12)",
         f"Bad Debt T12={ps(bd_v)} of GPR."),
        (scores['rev_trend'],
         f"Rental {'↑' if rv['rt']>=4 else '↓' if rv['rt']<=2 else '→'} {rv['rt']} | "
         f"Other {'↑' if rv['ot']>=4 else '↓' if rv['ot']<=2 else '→'} {rv['ot']}",
         f"Rental T12={ds(ri['T12'])} T3={ds(ri['T3'])} "
         f"({'+' if rv['rg_pct']>=0 else ''}{ps(rv['rg_pct'])}). "
         f"Other T12={ds(oi['T12'])} T3={ds(oi['T3'])} "
         f"({'+' if rv['og_pct']>=0 else ''}{ps(rv['og_pct'])}). "
         f"Composite={rv['composite']:.2f}→{rv['score']}."),
        (scores['tradeout'],
         f"Breadth {ps(rr['breadth_pct'])} | Spread {ps(rr['tradeout_spread'])}",
         f"{rr['n_leases_90d']} leases 90d. Breadth {ps(rr['breadth_pct'])} of "
         f"{rr['n_fps']} fps. New {ds(rr['avg_new_eff'])} vs Renewed "
         f"{ds(rr['avg_renewed_eff'])} "
         f"({'+' if rr['tradeout_spread']>=0 else ''}{ps(rr['tradeout_spread'])}). "
         f"Composite={tot['composite']:.2f}."),
        (scores['inplace'],
         f"In-place {ds(rr['avg_eff'])} vs Mkt {ds(rr['avg_market'])} "
         f"({'below' if ip>0 else 'above'} {ps(abs(ip))})",
         f"Avg in-place: {ds(rr['avg_eff'])}. Avg market: {ds(rr['avg_market'])}. "
         f"Spread {'+' if ip>=0 else ''}{ps(ip)}."),
    ]

    # ── Write only score (col D) + direction (col F) + notes (col G) ─────────
    # Weighted points (E), total (E15), band (E16) are template formulas — untouched.
    for i, (score, direction, notes) in enumerate(metric_data):
        r = 8 + i
        ws.cell(r, 4).value = score      # Score 1-5
        ws.cell(r, 6).value = direction  # Trend Direction
        ws.cell(r, 7).value = notes      # Notes / Evidence

    # ── Compute total + band locally for display (formulas handle this in Excel) ─
    total = sum(round(WEIGHTS[k] * scores[k] / 5)
                for k in ('noi','econ_occ','concessions','bad_debt',
                          'rev_trend','tradeout','inplace'))
    band  = get_band(total)
    log.write(f"  THS scores written  Score={total}  Band={band}\n")
    return total, band


def write_subscores(wb, scores, log):
    if 'Anchors & Notes' not in wb.sheetnames:
        log.write("  ⚠ Anchors & Notes not found — skipping sub-scores\n"); return
    an=wb['Anchors & Notes']; rv=scores['rev_trend_detail']; tot=scores['tradeout_detail']
    def ws2(row,col,val):
        cl=an.cell(row=row,column=col,value=val)
        if isinstance(val,(int,float)) and 1<=float(val)<=5 and float(val)==int(val):
            cl.fill=PatternFill("solid",fgColor=C_GREEN if int(val)>=4 else C_YELLOW if int(val)==3 else C_RED)
    ws2(33,5,rv['rt']); ws2(34,5,rv['rg']); ws2(35,5,rv['ot']); ws2(36,5,rv['og'])
    an.cell(37,5).value=round(rv['composite'],2)
    ws2(68,5,tot['breadth_score']); ws2(69,5,tot['spread_score'])
    an.cell(70,5).value=round(tot['composite'],2)
    log.write(f"  Sub-scores written (E33={rv['rt']} E34={rv['rg']} E35={rv['ot']} E36={rv['og']}  E68={tot['breadth_score']} E69={tot['spread_score']})\n")


# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Creative RE — UW Pipeline", page_icon="🏢", layout="centered")

st.markdown("""
<style>
    .block-container { max-width: 780px; }
    h1 { color: #1F3A5F; }
    .stDownloadButton > button { background-color: #1F3A5F; color: white; width: 100%; }
</style>
""", unsafe_allow_html=True)

st.title("🏢 Creative RE — UW Pipeline")
st.caption("Upload both files, fill in property details, review the column mapping, then run.")

st.divider()

# ── A. File uploads ───────────────────────────────────────────────────────────
st.markdown("**A. Upload Files**")
col1, col2 = st.columns(2)
with col1:
    ths_file = st.file_uploader("THS Template (T12 categorised)", type=["xlsx"],
                                help="Workbook with T12 and Rent Roll tabs. T12 must already be categorised.")
with col2:
    rr_file  = st.file_uploader("Raw Rent Roll", type=["xlsx", "xls"],
                                help="Rent roll export. Accepts .xlsx and legacy .xls files.")

st.divider()

# ── B. Property details ───────────────────────────────────────────────────────
st.markdown("**B. Property Details**")
c1, c2 = st.columns(2)
with c1:
    prop_name_input = st.text_input("Property name *", autocomplete="off",
                                    help="Required. Used in the scorecard header and output filename.")
with c2:
    address_input   = st.text_input("Property address *", autocomplete="off",
                                    help="Required. Used to look up floor plans online for unit mix.")

url_input = st.text_input("Property website / Apartments.com URL (optional)", autocomplete="off",
                          help="Paste URL to read floor plans directly from the page.")

st.divider()

# ── C. Rent Roll details ──────────────────────────────────────────────────────
st.markdown("**C. Rent Roll Details**")
rr_date_input = st.text_input("RR as-of date (e.g. 9/2/2025)", autocomplete="off",
                              help="Leave blank to auto-detect from RR file, filename, or T12.")

st.caption("Floor plan sizes — enter sq ft manually if URL lookup doesn't work (most reliable):")
fp_manual_input = st.text_area(
    "One per line — format: `Unit Mix: sqft` or `Unit Mix: min-max` or `Unit Mix: val1, val2`",
    placeholder="1 BR: 815, 900\n2 BR: 1100-1300\n3 BR: 1400",
    height=100,
)

st.divider()

# ── D. Column mapping ─────────────────────────────────────────────────────────
import hashlib as _hashlib

col_override = None

if rr_file:
    rr_bytes_raw = rr_file.getvalue()
    rr_hash = _hashlib.md5(rr_bytes_raw).hexdigest()

    if st.session_state.get('_rr_hash') != rr_hash:
        try:
            _detect_bytes = rr_bytes_raw
            if rr_file.name.lower().endswith('.xls'):
                _detect_bytes = _xls_to_xlsx_bytes(_detect_bytes)
            _wb_tmp = openpyxl.load_workbook(io.BytesIO(_detect_bytes), data_only=True)
            _ws_tmp = _wb_tmp['Report1'] if 'Report1' in _wb_tmp.sheetnames else _wb_tmp.worksheets[0]
            st.session_state['_rr_detection'] = _detect_columns(_ws_tmp)
            st.session_state['_rr_hash']      = rr_hash
            st.session_state['_rr_overrides'] = {}
        except Exception:
            st.session_state['_rr_detection'] = None

    detection = st.session_state.get('_rr_detection')

    if detection:
        st.markdown("**D. Column Mapping**")
        st.caption(
            f"Header row **{detection['header_row']}** · "
            f"Unit data column **{detection['unit_col']}** · "
            f"Format: **{'Multi-row charges' if detection['is_multirow'] else 'One row per unit'}**  "
            "— change any Override dropdown if a field is mapped to the wrong column."
        )

        # Only show fields that map to output template columns.
        # Status, Charge Code, Charge Amount are internal — excluded from the UI.
        # Unit Number shows the ACTUAL resolved column (after any left-shift).
        _FIELD_LABELS = {
            '_unit_col':   'Unit Number',       # uses resolved unit_col, not col_map['unit']
            'unit_type':   'Unit Type / Floor Plan',
            'sqft':        'Sq Ft',
            'tenant':      'Resident Name',
            'mkt_rent':    'Market Rent',
            'eff_rent':    'Effective Rent',
            'move_in':     'Move-In Date',
            'lease_from':  'Lease Start',
            'lease_to':    'Lease End',
        }

        _hdr_options = {c: txt for c, txt in detection['col_headers']}
        _sel_options = ['(not mapped)'] + [f"Col {c}: {txt}" for c, txt in detection['col_headers']]

        def _col_to_opt(c):
            if c is None: return '(not mapped)'
            return f"Col {c}: {_hdr_options.get(c, str(c))}"
        def _opt_to_col(s):
            if s == '(not mapped)': return None
            try: return int(s.split(':')[0].replace('Col', '').strip())
            except: return None

        overrides = st.session_state.get('_rr_overrides', {})
        new_overrides = {}

        # Table header
        h1, h2, h3 = st.columns([2, 2.5, 3])
        h1.markdown("**Metric**")
        h2.markdown("**Auto-Detected Column**")
        h3.markdown("**Override**")
        st.markdown('<hr style="margin:2px 0 8px 0">', unsafe_allow_html=True)

        for field, label in _FIELD_LABELS.items():
            # Unit Number uses the resolved unit_col; all others use col_map
            if field == '_unit_col':
                detected_col = detection['unit_col']
                current_col  = overrides.get('_unit_col', detected_col)
            else:
                detected_col = detection['col_map'].get(field)
                current_col  = overrides.get(field, detected_col)

            current_opt = _col_to_opt(current_col)
            r1, r2, r3  = st.columns([2, 2.5, 3])
            r1.write(label)
            r2.write(_col_to_opt(detected_col))
            sel = r3.selectbox("", _sel_options,
                               index=_sel_options.index(current_opt) if current_opt in _sel_options else 0,
                               key=f"_cm_{field}", label_visibility="collapsed")
            chosen = _opt_to_col(sel)
            if chosen != detected_col:
                new_overrides[field] = chosen

        st.session_state['_rr_overrides'] = new_overrides

        # Build col_override for parse_rent_roll
        final_col_map = detection['col_map'].copy()
        # Apply field overrides (excluding _unit_col which is handled separately)
        final_col_map.update({k: v for k, v in new_overrides.items() if not k.startswith('_')})
        final_unit_col = new_overrides.get('_unit_col', detection['unit_col'])

        col_override = dict(
            header_row  = detection['header_row'],
            col_map     = final_col_map,
            col_headers = detection['col_headers'],
            data_start  = detection['data_start'],
            data_end    = detection['data_end'],
            unit_col    = final_unit_col,
            is_multirow = detection['is_multirow'],
        )

        st.divider()
    else:
        st.divider()

_ready = (ths_file is not None and rr_file is not None
          and prop_name_input.strip() and address_input.strip())
run = st.button("▶  Run UW Pipeline", type="primary", disabled=not _ready)

if not _ready and (ths_file or rr_file):
    missing = []
    if not prop_name_input.strip(): missing.append("property name")
    if not address_input.strip():   missing.append("property address")
    if missing:
        st.warning(f"Please enter the {' and '.join(missing)} before running.")

if ths_file and rr_file and prop_name_input.strip() and address_input.strip() and run:
    log_buf = io.StringIO()

    try:
        # ── Convert .xls → .xlsx if needed ───────────────────────────────
        rr_bytes = rr_file.getvalue()
        if rr_file.name.lower().endswith('.xls'):
            rr_bytes = _xls_to_xlsx_bytes(rr_bytes)
            log_buf.write("  Converted .xls → .xlsx\n")

        # ── Property name (required — typed by user) ──────────────────────
        prop_name = prop_name_input.strip()

        # ── RR date: user-typed > RR file content > filename > T12 date > today ──
        rr_date = None
        rr_date_source = ''

        # 1. User explicitly typed a date
        if rr_date_input.strip():
            rr_date = _parse_date_from_string(rr_date_input.strip())
            if rr_date:
                rr_date_source = 'user input'
            else:
                st.warning(f"Could not parse '{rr_date_input.strip()}' as a date — auto-detecting.")

        # 2. Scan the top of the RR file itself for a date
        if rr_date is None:
            _log_tmp = io.StringIO()
            rr_date = _extract_date_from_rr_file(rr_bytes, _log_tmp)
            if rr_date:
                rr_date_source = f'RR file ({_log_tmp.getvalue().strip()})'

        # 3. Check the RR filename for a date pattern
        if rr_date is None:
            m = re.search(r'(\d{4})[.\-](\d{2})[.\-](\d{2})', rr_file.name)
            if not m:
                m = re.search(r'(\d{2})[.\-](\d{2})[.\-](\d{4})', rr_file.name)
                if m:
                    try:
                        rr_date = date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
                        rr_date_source = 'filename (MM-DD-YYYY)'
                    except: pass
            else:
                try:
                    rr_date = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                    rr_date_source = 'filename (YYYY-MM-DD)'
                except: pass

        # 4. Fall back to the last month in the T12
        if rr_date is None:
            _log_tmp = io.StringIO()
            rr_date = _extract_date_from_t12(ths_file.getvalue(), _log_tmp)
            if rr_date:
                rr_date_source = f'T12 last month ({_log_tmp.getvalue().strip()})'

        # 5. Last resort — today
        if rr_date is None:
            rr_date = date.today()
            rr_date_source = 'today (fallback)'
            st.warning(f"Could not detect RR date — using today ({rr_date}). Set it manually if incorrect.")

        st.caption(f"RR date: **{rr_date.strftime('%m/%d/%Y')}** — detected from {rr_date_source}")

        today_str  = rr_date.strftime("%m/%d/%Y")
        # Output filename: Creative RE_THS_Property Name_MM.DD.YYYY (date of generation)
        gen_tag   = date.today().strftime("%m.%d.%Y")
        out_name  = f"Creative RE_THS_{prop_name}_{gen_tag}.xlsx"

        log_buf.write(f"Property : {prop_name}\n")
        log_buf.write(f"RR date  : {rr_date}\n\n")

        with st.status("Running pipeline ...", expanded=True) as status:
            # Step 1 — Floor plan lookup (manual override → URL → search)
            st.write("Resolving floor plans ...")
            log_buf.write("[1] Floor plan lookup\n")

            floor_plans = []

            # 1a — Manual textarea (most reliable — always try first)
            # Supported formats per line (mix freely):
            #   1 BR: 815                     → single exact sq ft
            #   1 BR: 800-950                 → range
            #   1 BR: 800, 900, 1200          → multiple exact sq fts
            #   1 BR: 800-950, 1200           → range + exact value
            if fp_manual_input.strip():
                def _normalise_label(raw):
                    lbl = raw.strip().title()
                    for std in ['Studio','1 BR','2 BR','3 BR','4 BR','5 BR']:
                        if re.search(
                            std.replace(' ', r'[\s\-]?').replace('BR', r'br', 1),
                            lbl, re.IGNORECASE):
                            return std
                    return lbl

                def _parse_fp_values(label, values_str):
                    """Parse 'val1, val2, range1-range2, ...' into (lo,hi,label) tuples."""
                    tuples = []
                    for part in re.split(r',\s*', values_str.strip()):
                        part = part.strip()
                        # Range: "800-950" or "800 – 950"
                        rm = re.match(r'^(\d{3,5})\s*[-–to]+\s*(\d{3,5})$', part, re.IGNORECASE)
                        if rm:
                            tuples.append((int(rm.group(1)), int(rm.group(2)), label))
                        else:
                            # Single value
                            sm = re.match(r'^(\d{3,5})$', part)
                            if sm:
                                v = int(sm.group(1))
                                tuples.append((v, v, label))
                    return tuples

                for line in fp_manual_input.strip().splitlines():
                    line = line.strip()
                    if not line or ':' not in line: continue
                    label_part, _, values_part = line.partition(':')
                    label = _normalise_label(label_part)
                    floor_plans.extend(_parse_fp_values(label, values_part))

                if floor_plans:
                    log_buf.write(f"  Manual floor plans: {floor_plans}\n")

            # 1b — URL or search (only if manual not provided)
            if not floor_plans:
                floor_plans = _lookup_unit_mix_online(
                    prop_name, address_input.strip(), url_input.strip(), log_buf)

            if floor_plans:
                log_buf.write(f"  {len(floor_plans)} floor plan range(s) active\n")
            else:
                log_buf.write("  No floor plan data — unit mix from type code only\n")

            # Step 2 — Parse RR (one-liner internally; used to fill Rent Roll tab)
            st.write("Parsing rent roll ...")
            log_buf.write("\n[2] Parsing rent roll\n")
            units, src_totals, source_totals = parse_rent_roll(
                rr_bytes, rr_date, log_buf, floor_plans, col_override=col_override)

            # Step 3 — Open THS + write Rent Roll tab
            st.write("Writing Rent Roll tab ...")
            log_buf.write("\n[3] Writing Rent Roll tab\n")
            wb = openpyxl.load_workbook(io.BytesIO(ths_file.getvalue()))
            if 'Rent Roll' not in wb.sheetnames:
                raise ValueError("THS template has no 'Rent Roll' sheet. Please check the file.")
            if 'T12' not in wb.sheetnames:
                raise ValueError("THS template has no 'T12' sheet. Please check the file.")
            write_rr_tab(wb, units, src_totals, source_totals, rr_date, prop_name, log_buf)

            # Write property name to T12 tab C4
            if 'T12' in wb.sheetnames:
                wb['T12'].cell(4, 3).value = prop_name
                log_buf.write(f"  T12 C4 ← '{prop_name}'\n")

            # Step 4 — Read T12 metrics
            st.write("Reading T12 metrics ...")
            log_buf.write("\n[4] Reading T12 metrics\n")
            t12_metrics = read_t12_metrics(wb, log_buf)

            # Step 5 — RR metrics
            st.write("Computing RR metrics ...")
            log_buf.write("\n[5] RR metrics\n")
            rr_metrics = read_rr_metrics(units, src_totals, rr_date, log_buf)

            # Step 6 — Score
            st.write("Scoring ...")
            log_buf.write("\n[6] Scoring\n")
            scores = score_all(t12_metrics, rr_metrics, log_buf)

            # Step 7 — Build scorecard sheet
            st.write("Building scorecard sheet ...")
            log_buf.write("\n[7] Building THS sheet\n")
            total_score, band = build_ths_sheet(wb, t12_metrics, rr_metrics, scores, today_str, prop_name, log_buf)

            # Step 8 — Sub-scores
            log_buf.write("\n[8] Writing sub-scores\n")
            write_subscores(wb, scores, log_buf)

            # Step 9 — Save to bytes
            log_buf.write("\n[9] Saving\n")
            out_buf = io.BytesIO()
            wb.save(out_buf)
            out_buf.seek(0)
            out_bytes = out_buf.getvalue()

            # Verify ZIP
            try:
                with zipfile.ZipFile(io.BytesIO(out_bytes)) as z:
                    bad = z.testzip()
                    if bad: raise zipfile.BadZipFile(bad)
                log_buf.write("  ✓ ZIP integrity verified\n")
            except zipfile.BadZipFile as e:
                raise RuntimeError(f"Output ZIP integrity failed: {e}")

            status.update(label="✅ Pipeline complete!", state="complete")

        # ── Store results in session state so they survive reruns ─────────
        st.session_state['_result'] = dict(
            out_bytes   = out_bytes,
            out_name    = out_name,
            total_score = total_score,
            band        = band,
            n_units     = src_totals['n_units'],
            log         = log_buf.getvalue(),
        )

    except Exception as e:
        st.error(f"**Error:** {e}")
        with st.expander("Pipeline log"):
            st.code(log_buf.getvalue(), language=None)
        st.session_state.pop('_result', None)   # clear stale result on error

# ── Show results whenever they exist in session state ─────────────────────────
if '_result' in st.session_state:
    res = st.session_state['_result']
    st.divider()
    bcol1, bcol2, bcol3 = st.columns(3)
    bcol1.metric("Score",   res['total_score'])
    bcol2.metric("Band",    res['band'])
    bcol3.metric("Units",   res['n_units'])

    st.download_button(
        label=f"⬇  Download {res['out_name']}",
        data=res['out_bytes'],
        file_name=res['out_name'],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    with st.expander("Pipeline log"):
        st.code(res['log'], language=None)

elif ths_file is None or rr_file is None:
    st.info("Upload both files above to enable the Run button.")
